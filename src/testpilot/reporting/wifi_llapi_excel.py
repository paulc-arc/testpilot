"""wifi_llapi Excel report helpers.

This module keeps report layout compatible with the reference
`Wifi_LLAPI` worksheet and supports:
1) building a template report from source workbook,
2) creating dated run reports from template,
3) filling test command and result columns by case source row.
"""

from __future__ import annotations

from dataclasses import dataclass
from datetime import date
import json
from pathlib import Path
import re
import shutil
from typing import Iterable

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.utils import column_index_from_string

DEFAULT_SHEET_NAME = "Wifi_LLAPI"
DEFAULT_TEMPLATE_NAME = "wifi_llapi_template.xlsx"
DEFAULT_CLEAR_COLUMNS = (
    "G",   # Test steps
    "H",   # Driver-level verified command output
    "R",   # Tester (BCM)
    "S",   # BCM 5g result
    "T",   # BCM 6g result
    "U",   # BCM 2.4g result
    "V",   # BCM comment
    "W",   # BCM issue type/internal notes
    "X",   # ARC 5g result
    "Y",   # ARC 6g result
    "Z",   # ARC 2.4g result
    "AA",  # ARC tester
    "AB",  # ARC comment
)
DATA_START_ROW = 4
EMPTY_STREAK_STOP = 200
MAX_SCAN_ROWS = 5000


@dataclass(slots=True)
class ClearRules:
    """Template clear behavior for wifi_llapi sheet."""

    clear_columns: tuple[str, ...] = DEFAULT_CLEAR_COLUMNS
    data_start_row: int = DATA_START_ROW
    empty_streak_stop: int = EMPTY_STREAK_STOP
    max_scan_rows: int = MAX_SCAN_ROWS


@dataclass(slots=True)
class TemplateBuildResult:
    template_path: Path
    sheet_name: str
    total_case_rows: int
    cleared_columns: tuple[str, ...]
    source_workbook: Path
    source_sheet: str


@dataclass(slots=True)
class ReportMeta:
    run_date: date
    dut_fw_ver: str
    source_excel: str
    sheet_name: str = DEFAULT_SHEET_NAME


@dataclass(slots=True)
class WifiLlapiCaseResult:
    case_id: str
    source_row: int
    executed_test_command: str
    command_output: str
    result_5g: str
    result_6g: str
    result_24g: str
    comment: str = ""
    tester: str = "testpilot"


def normalize_text(text: str | None) -> str:
    if text is None:
        return ""
    cleaned = re.sub(r"[\u200b\u200c\u200d\ufeff]", "", str(text))
    cleaned = re.sub(r"\s+", " ", cleaned)
    return cleaned.strip()


def sanitize_fw_version(text: str) -> str:
    cleaned = re.sub(r"[^A-Za-z0-9._-]+", "-", text.strip())
    cleaned = cleaned.strip("-._")
    return cleaned or "DUT-FW-VER"


def generate_report_filename(run_date: date, dut_fw_ver: str) -> str:
    """Generate user-required report filename format.

    Format: YYYYMMDD_<DUT-FW-VER>_wifi_LLAPI.xlsx
    """
    return f"{run_date:%Y%m%d}_{sanitize_fw_version(dut_fw_ver)}_wifi_LLAPI.xlsx"


def _get_sheet(wb, preferred_name: str):
    if preferred_name in wb.sheetnames:
        return wb[preferred_name]
    lower_map = {n.lower(): n for n in wb.sheetnames}
    if preferred_name.lower() in lower_map:
        return wb[lower_map[preferred_name.lower()]]
    raise KeyError(f"sheet not found: {preferred_name}, available={wb.sheetnames}")


def _iter_case_rows(
    ws,
    data_start_row: int,
    empty_streak_stop: int,
    max_scan_rows: int,
) -> list[int]:
    """Detect continuous test-case rows by Parameter Name column (C)."""
    rows: list[int] = []
    empty_streak = 0
    row = data_start_row
    max_row = min(ws.max_row, data_start_row + max_scan_rows)
    while row <= max_row:
        value = ws.cell(row=row, column=3).value
        if value is None or str(value).strip() == "":
            empty_streak += 1
            if empty_streak >= empty_streak_stop:
                break
        else:
            rows.append(row)
            empty_streak = 0
        row += 1
    return rows


def _to_col_idx(col: str | int) -> int:
    if isinstance(col, int):
        return col
    return column_index_from_string(col)


def build_template_from_source(
    source_xlsx: Path | str,
    out_template_xlsx: Path | str,
    *,
    sheet_name: str = DEFAULT_SHEET_NAME,
    clear_rules: ClearRules | None = None,
) -> TemplateBuildResult:
    """Extract wifi_LLAPI sheet as template and clear result/test command fields."""
    rules = clear_rules or ClearRules()
    source = Path(source_xlsx)
    out_path = Path(out_template_xlsx)
    out_path.parent.mkdir(parents=True, exist_ok=True)

    wb = load_workbook(source)
    ws = _get_sheet(wb, sheet_name)
    actual_sheet_name = ws.title

    # Keep only target sheet to match "extract this page".
    for name in list(wb.sheetnames):
        if name != actual_sheet_name:
            wb.remove(wb[name])
    wb.active = 0
    ws = wb[actual_sheet_name]

    case_rows = _iter_case_rows(
        ws,
        rules.data_start_row,
        rules.empty_streak_stop,
        rules.max_scan_rows,
    )
    clear_col_indices = [_to_col_idx(c) for c in rules.clear_columns]
    for row in case_rows:
        for col_idx in clear_col_indices:
            cell = ws.cell(row=row, column=col_idx)
            if isinstance(cell, MergedCell):
                continue
            cell.value = None

    wb.save(out_path)
    wb.close()

    return TemplateBuildResult(
        template_path=out_path,
        sheet_name=actual_sheet_name,
        total_case_rows=len(case_rows),
        cleared_columns=rules.clear_columns,
        source_workbook=source,
        source_sheet=actual_sheet_name,
    )


def write_template_manifest(
    manifest_path: Path | str,
    build_result: TemplateBuildResult,
) -> Path:
    payload = {
        "template_path": str(build_result.template_path),
        "sheet_name": build_result.sheet_name,
        "total_case_rows": build_result.total_case_rows,
        "cleared_columns": list(build_result.cleared_columns),
        "source_workbook": str(build_result.source_workbook),
        "source_sheet": build_result.source_sheet,
    }
    out = Path(manifest_path)
    out.parent.mkdir(parents=True, exist_ok=True)
    out.write_text(json.dumps(payload, indent=2, ensure_ascii=False), encoding="utf-8")
    return out


def create_run_report_from_template(
    template_xlsx: Path | str,
    out_report_xlsx: Path | str,
) -> Path:
    src = Path(template_xlsx)
    dst = Path(out_report_xlsx)
    dst.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(src, dst)
    return dst


def fill_case_results(
    report_xlsx: Path | str,
    case_results: Iterable[WifiLlapiCaseResult],
    *,
    sheet_name: str = DEFAULT_SHEET_NAME,
) -> Path:
    """Batch fill test command and result columns by source row."""
    path = Path(report_xlsx)
    wb = load_workbook(path)
    ws = _get_sheet(wb, sheet_name)

    for item in case_results:
        if item.source_row <= 0:
            continue
        row = item.source_row
        ws[f"G{row}"] = item.executed_test_command
        ws[f"H{row}"] = item.command_output
        ws[f"R{row}"] = item.tester
        ws[f"S{row}"] = item.result_5g
        ws[f"T{row}"] = item.result_6g
        ws[f"U{row}"] = item.result_24g
        ws[f"V{row}"] = item.comment

    wb.save(path)
    wb.close()
    return path


def finalize_report_metadata(
    report_xlsx: Path | str,
    meta: ReportMeta,
) -> Path:
    """Store report metadata in a hidden _meta sheet."""
    path = Path(report_xlsx)
    wb = load_workbook(path)
    if "_meta" in wb.sheetnames:
        ws = wb["_meta"]
        ws.delete_rows(1, ws.max_row)
    else:
        ws = wb.create_sheet("_meta")
        ws.sheet_state = "hidden"

    ws["A1"] = "run_date"
    ws["B1"] = meta.run_date.isoformat()
    ws["A2"] = "dut_fw_ver"
    ws["B2"] = meta.dut_fw_ver
    ws["A3"] = "source_excel"
    ws["B3"] = meta.source_excel
    ws["A4"] = "sheet_name"
    ws["B4"] = meta.sheet_name

    wb.save(path)
    wb.close()
    return path


def ensure_template_report(
    source_xlsx: Path | str,
    template_path: Path | str,
    *,
    manifest_path: Path | str | None = None,
    sheet_name: str = DEFAULT_SHEET_NAME,
) -> TemplateBuildResult:
    """Build template and optional manifest."""
    result = build_template_from_source(
        source_xlsx=source_xlsx,
        out_template_xlsx=template_path,
        sheet_name=sheet_name,
    )
    if manifest_path is not None:
        write_template_manifest(manifest_path, result)
    return result


def read_source_rows(
    source_xlsx: Path | str,
    *,
    sheet_name: str = DEFAULT_SHEET_NAME,
) -> dict[int, dict[str, str]]:
    """Read source object/api by row from Wifi_LLAPI sheet."""
    source = Path(source_xlsx)
    wb = load_workbook(source, read_only=True, data_only=False)
    ws = _get_sheet(wb, sheet_name)
    rows = _iter_case_rows(ws, DATA_START_ROW, EMPTY_STREAK_STOP, MAX_SCAN_ROWS)
    out: dict[int, dict[str, str]] = {}
    current_object = ""
    for row in rows:
        obj = ws.cell(row=row, column=1).value
        api = ws.cell(row=row, column=3).value
        obj_norm = normalize_text(obj)
        if obj_norm:
            current_object = obj_norm
        out[row] = {
            "object": current_object,
            "api": normalize_text(api),
        }
    wb.close()
    return out


def collect_alignment_issues(
    cases: Iterable[dict],
    source_xlsx: Path | str,
    *,
    sheet_name: str = DEFAULT_SHEET_NAME,
) -> list[dict[str, str | int]]:
    """Compare case source.row/object/api against source Excel sheet."""
    source_rows = read_source_rows(source_xlsx, sheet_name=sheet_name)
    issues: list[dict[str, str | int]] = []
    for case in cases:
        source = case.get("source", {}) if isinstance(case.get("source"), dict) else {}
        case_id = str(case.get("id", "?"))
        try:
            row = int(source.get("row", 0))
        except (TypeError, ValueError):
            row = 0

        expected_object = normalize_text(source.get("object"))
        expected_api = normalize_text(source.get("api"))

        if row not in source_rows:
            issues.append(
                {
                    "case_id": case_id,
                    "source_row": row,
                    "issue": "row_not_found",
                    "expected_object": expected_object,
                    "expected_api": expected_api,
                    "sheet_object": "",
                    "sheet_api": "",
                }
            )
            continue

        actual_object = source_rows[row]["object"]
        actual_api = source_rows[row]["api"]
        if expected_object != actual_object or expected_api != actual_api:
            issues.append(
                {
                    "case_id": case_id,
                    "source_row": row,
                    "issue": "object_or_api_mismatch",
                    "expected_object": expected_object,
                    "expected_api": expected_api,
                    "sheet_object": actual_object,
                    "sheet_api": actual_api,
                }
            )
    return issues
