from datetime import date
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

from testpilot.reporting.wifi_llapi_excel import (
    COMMENT_HEADER,
    DEFAULT_CLEAR_COLUMNS,
    DEFAULT_TEMPLATE_MAX_COLUMN,
    TemplateValidationError,
    WifiLlapiCaseResult,
    build_template_from_source,
    collect_alignment_issues,
    create_run_report_from_template,
    fill_case_results,
    generate_report_filename,
    normalize_command_block,
    read_wifi_llapi_template_objects,
    sanitize_report_output,
    validate_wifi_llapi_report_template,
    write_summary_sheet,
    _truncate_comment,
)
from testpilot.schema.case_schema import load_case


def _create_source_xlsx(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Wifi_LLAPI"
    wb.create_sheet("OtherSheet")

    ws["A2"] = "Object"
    ws["C2"] = "Parameter Name"
    ws["G2"] = "Test steps"
    ws["H2"] = "Command Output"
    ws["I2"] = "ARC 4.0.3 Test Result\nWiFi 5g"
    ws["J2"] = "ARC 4.0.3 Test Result\nWiFi 6g"
    ws["K2"] = "ARC 4.0.3 Test Result\nWiFi 2.4g"
    ws["L2"] = "Tester"
    ws["M2"] = "Legacy Extra Band"
    ws["S1"] = "BCM"
    ws["S2"] = "BCM v4.0.3 Test Result WiFi 5g"
    ws["X1"] = "ARC"
    ws["X2"] = "ARC 4.0.3 Test Result WiFi 5g"
    ws.merge_cells("I1:Q1")
    ws.merge_cells("L2:M2")
    ws.column_dimensions["S"].width = 18
    ws.column_dimensions["AB"].width = 24

    ws["A4"] = "WiFi.AccessPoint.{i}."
    ws["C4"] = "kickStation()"
    ws["G4"] = "old-step"
    ws["H4"] = "old-output"
    ws["I4"] = "Pass"
    ws["J4"] = "Pass"
    ws["K4"] = "Pass"
    ws["L4"] = "old-tester"

    ws["A5"] = "WiFi.Radio.{i}."
    ws["C5"] = "scan()"
    ws["G5"] = "old-step2"
    ws["H5"] = "old-output2"
    ws["I5"] = "Fail"
    ws["J5"] = "Fail"
    ws["K5"] = "Fail"
    ws["L5"] = "old-tester2"

    wb.save(path)
    wb.close()


def test_build_template_from_source(tmp_path: Path):
    source = tmp_path / "source.xlsx"
    template = tmp_path / "wifi_llapi_template.xlsx"
    _create_source_xlsx(source)

    result = build_template_from_source(source, template)
    assert result.total_case_rows == 2
    assert result.sheet_name == "Wifi_LLAPI"
    assert DEFAULT_TEMPLATE_MAX_COLUMN == "M"
    assert result.cleared_columns == DEFAULT_CLEAR_COLUMNS
    assert "M" in result.cleared_columns

    wb = load_workbook(template)
    assert wb.sheetnames == ["Wifi_LLAPI"]
    ws = wb["Wifi_LLAPI"]
    assert ws.max_column == 16
    assert get_column_letter(ws.max_column) == "P"
    assert all(r.max_col <= 13 for r in ws.merged_cells.ranges)
    assert ws["I2"].value == "Result"
    assert ws["L2"].value == "Tester"
    assert ws["I3"].value == "WiFi 5G"
    assert ws["J3"].value == "WiFi 6G"
    assert ws["K3"].value == "WiFi 2.4G"
    assert ws["L3"].value == "Tester"
    assert ws["M3"].value == COMMENT_HEADER
    assert ws["N3"].value == "Summary Bucket WiFi 5G"
    assert ws["O3"].value == "Summary Bucket WiFi 6G"
    assert ws["P3"].value == "Summary Bucket WiFi 2.4G"
    assert ws.column_dimensions["N"].hidden is True
    assert ws.column_dimensions["O"].hidden is True
    assert ws.column_dimensions["P"].hidden is True
    assert "S" not in ws.column_dimensions
    assert "AB" not in ws.column_dimensions
    assert ws["C4"].value == "kickStation()"
    assert ws["C5"].value == "scan()"
    for cell in ("G4", "H4", "I4", "J4", "K4", "L4", "M4", "G5", "H5", "I5", "J5", "K5", "L5", "M5"):
        assert ws[cell].value is None
    wb.close()


def test_build_template_from_source_preserves_existing_summary_contract(tmp_path: Path):
    source = tmp_path / "source.xlsx"
    template = tmp_path / "wifi_llapi_template.xlsx"
    _create_source_xlsx(source)
    _create_template_with_summary(template)

    wb_before = load_workbook(template, data_only=False)
    before = (
        wb_before["Summary"]["C1"].value,
        wb_before["Summary"]["F3"].value,
        wb_before["Summary"]["J3"].number_format,
        [str(rng) for rng in wb_before["Summary"].merged_cells.ranges],
    )
    wb_before.close()

    build_template_from_source(source, template)

    wb = load_workbook(template, data_only=False)
    assert wb.sheetnames == ["Summary", "Wifi_LLAPI"]
    ws_summary = wb["Summary"]
    after = (
        ws_summary["C1"].value,
        ws_summary["F3"].value,
        ws_summary["J3"].number_format,
        [str(rng) for rng in ws_summary.merged_cells.ranges],
    )
    assert after == before
    ws_wifi = wb["Wifi_LLAPI"]
    assert ws_wifi["N3"].value == "Summary Bucket WiFi 5G"
    assert ws_wifi["O3"].value == "Summary Bucket WiFi 6G"
    assert ws_wifi["P3"].value == "Summary Bucket WiFi 2.4G"
    assert ws_wifi.column_dimensions["N"].hidden is True
    assert ws_wifi.column_dimensions["O"].hidden is True
    assert ws_wifi.column_dimensions["P"].hidden is True
    wb.close()


def test_fill_case_results(tmp_path: Path):
    source = tmp_path / "source.xlsx"
    template = tmp_path / "wifi_llapi_template.xlsx"
    report = tmp_path / "20260304_BGW720-B0-403_wifi_LLAPI.xlsx"
    _create_source_xlsx(source)
    build_template_from_source(source, template)
    create_run_report_from_template(template, report)

    fill_case_results(
        report,
        [
            WifiLlapiCaseResult(
                case_id="wifi-llapi-D004-kickstation",
                source_row=4,
                executed_test_command="ubus-cli ...kickStation...\n\nwl assoclist",
                command_output="root@prplOS:/# wl assoclist\nassoclist empty\n>\n",
                result_5g="Pass",
                result_6g="Pass",
                result_24g="Pass",
                comment="verified via DUT and STA logs",
            )
        ],
    )

    wb = load_workbook(report)
    ws = wb["Wifi_LLAPI"]
    assert ws["G4"].value == "ubus-cli ...kickStation...\nwl assoclist"
    assert ws["H4"].value == "wl assoclist\nassoclist empty"
    assert ws["I4"].value == "Pass"
    assert ws["J4"].value == "Pass"
    assert ws["K4"].value == "Pass"
    assert ws["L4"].value == "testpilot"
    assert ws["M4"].value == "verified via DUT and STA logs"
    wb.close()


def test_fill_case_results_comment_truncates_and_blank_stays_empty(tmp_path: Path):
    source = tmp_path / "source.xlsx"
    template = tmp_path / "wifi_llapi_template.xlsx"
    report = tmp_path / "20260304_BGW720-B0-403_wifi_LLAPI.xlsx"
    _create_source_xlsx(source)
    build_template_from_source(source, template)
    create_run_report_from_template(template, report)

    fill_case_results(
        report,
        [
            WifiLlapiCaseResult(
                case_id="wifi-llapi-D004-kickstation",
                source_row=4,
                executed_test_command="cmd",
                command_output="out",
                result_5g="Pass",
                result_6g="Pass",
                result_24g="Pass",
                comment="x" * 201,
            ),
            WifiLlapiCaseResult(
                case_id="wifi-llapi-D005-scan",
                source_row=5,
                executed_test_command="cmd2",
                command_output="out2",
                result_5g="Fail",
                result_6g="Fail",
                result_24g="Fail",
                comment="",
            ),
        ],
    )

    wb = load_workbook(report)
    ws = wb["Wifi_LLAPI"]
    assert ws["M4"].value == ("x" * 197) + "..."
    assert len(ws["M4"].value) == 200
    assert _truncate_comment("") == ""
    assert _truncate_comment(None) == ""
    assert ws["M5"].value in (None, "")
    wb.close()


def test_fill_case_results_with_merged_row(tmp_path: Path):
    source = tmp_path / "source.xlsx"
    template = tmp_path / "wifi_llapi_template.xlsx"
    report = tmp_path / "20260304_BGW720-B0-403_wifi_LLAPI.xlsx"
    _create_source_xlsx(source)

    wb = load_workbook(source)
    ws = wb["Wifi_LLAPI"]
    ws.merge_cells("G4:G5")
    ws.merge_cells("H4:H5")
    ws.merge_cells("I4:I5")
    ws.merge_cells("J4:J5")
    ws.merge_cells("K4:K5")
    ws.merge_cells("L4:L5")
    ws.merge_cells("M4:M5")
    wb.save(source)
    wb.close()

    build_template_from_source(source, template)
    create_run_report_from_template(template, report)

    fill_case_results(
        report,
        [
            WifiLlapiCaseResult(
                case_id="wifi-llapi-D005-merged",
                source_row=5,
                executed_test_command="cmd-from-merged-row",
                command_output="out-from-merged-row",
                result_5g="Fail",
                result_6g="N/A",
                result_24g="Pass",
            )
        ],
    )

    wb = load_workbook(report)
    ws = wb["Wifi_LLAPI"]
    assert ws["G4"].value == "cmd-from-merged-row"
    assert ws["H4"].value == "out-from-merged-row"
    assert ws["I4"].value == "Fail"
    assert ws["J4"].value == "N/A"
    assert ws["K4"].value == "Pass"
    assert ws["L4"].value == "testpilot"
    assert ws["M4"].value in (None, "")
    wb.close()


def test_generate_report_filename():
    assert (
        generate_report_filename(date(2026, 3, 4), "BGW720-B0-403")
        == "20260304_BGW720-B0-403_wifi_LLAPI.xlsx"
    )
    assert (
        generate_report_filename(date(2026, 3, 4), "BGW720 B0 403")
        == "20260304_BGW720-B0-403_wifi_LLAPI.xlsx"
    )
    assert (
        generate_report_filename(
            date(2026, 3, 4),
            "BGW720 B0 403",
            unique_suffix="20260304T101112123456",
        )
        == "20260304_BGW720-B0-403_wifi_LLAPI_20260304T101112123456.xlsx"
    )


def test_create_run_report_from_template_preserves_existing_report(tmp_path: Path):
    source = tmp_path / "source.xlsx"
    template = tmp_path / "wifi_llapi_template.xlsx"
    report = tmp_path / "20260304_BGW720-B0-403_wifi_LLAPI.xlsx"
    _create_source_xlsx(source)
    build_template_from_source(source, template)

    first = create_run_report_from_template(template, report)
    second = create_run_report_from_template(template, report)

    assert first == report
    assert second != report
    assert second.name == "20260304_BGW720-B0-403_wifi_LLAPI_01.xlsx"
    assert first.is_file()
    assert second.is_file()


def test_report_output_and_command_helpers_clean_noise():
    assert normalize_command_block("cmd1\n\n cmd2 \n") == "cmd1\ncmd2"
    assert (
        sanitize_report_output(
            "root@prplOS:/# wl status\nstatus ok\n>\n"
        )
        == "wl status\nstatus ok"
    )


def test_collect_alignment_issues(tmp_path: Path):
    source = tmp_path / "source.xlsx"
    _create_source_xlsx(source)
    cases = [
        {
            "id": "wifi-llapi-kickstation",
            "source": {
                "row": 4,
                "object": "WiFi.AccessPoint.{i}.",
                "api": "kickStation()",
            },
        },
        {
            "id": "wifi-llapi-mismatch",
            "source": {
                "row": 5,
                "object": "WiFi.Radio.{i}.",
                "api": "wrongApi()",
            },
        },
    ]
    issues = collect_alignment_issues(cases, source)
    assert len(issues) == 1
    assert issues[0]["case_id"] == "wifi-llapi-mismatch"
    assert issues[0]["issue"] == "object_or_api_mismatch"


def test_collect_alignment_issues_on_repo_template_case():
    repo_root = Path(__file__).resolve().parents[3]
    case = load_case(
        repo_root / "plugins/wifi_llapi/cases/D018_downlinkshortguard.yaml"
    )
    template = repo_root / "plugins/wifi_llapi/reports/templates/wifi_llapi_template.xlsx"

    issues = collect_alignment_issues([case], template)

    assert issues == []


def test_repo_template_summary_is_formula_driven_and_styled():
    repo_root = Path(__file__).resolve().parents[3]
    template = repo_root / "plugins/wifi_llapi/reports/templates/wifi_llapi_template.xlsx"

    assert validate_wifi_llapi_report_template(template) == {
        "summary_sheet": "Summary",
        "wifi_sheet": "Wifi_LLAPI",
    }

    wb = load_workbook(template, data_only=False)
    ws = wb["Summary"]

    assert "A3:A8" in {str(rng) for rng in ws.merged_cells.ranges}
    assert "C1:J1" in {str(rng) for rng in ws.merged_cells.ranges}
    assert ws["C1"].value == "Detailed Statistics by Object"
    assert ws["C1"].fill.fill_type == "solid"

    assert ws["C3"].value == '=COUNTIF(Wifi_LLAPI!$A:$A,$B3&"*")'
    assert ws["E3"].value == (
        '=COUNTIFS(Wifi_LLAPI!$A:$A,$B3&"*",Wifi_LLAPI!$N:$N,E$2)'
    )
    assert ws["F3"].value == (
        '=COUNTIFS(Wifi_LLAPI!$A:$A,$B3&"*",Wifi_LLAPI!$N:$N,F$2)'
    )
    assert ws["G3"].value == (
        '=COUNTIFS(Wifi_LLAPI!$A:$A,$B3&"*",Wifi_LLAPI!$N:$N,G$2)'
    )
    assert ws["H3"].value == (
        '=COUNTIFS(Wifi_LLAPI!$A:$A,$B3&"*",Wifi_LLAPI!$N:$N,H$2)'
        '+COUNTIFS(Wifi_LLAPI!$A:$A,$B3&"*",Wifi_LLAPI!$E:$E,'
        '"Not Supported",Wifi_LLAPI!$N:$N,"")'
    )
    assert ws["I3"].value == (
        '=COUNTIFS(Wifi_LLAPI!$A:$A,$B3&"*",Wifi_LLAPI!$N:$N,I$2)'
        '+COUNTIFS(Wifi_LLAPI!$A:$A,$B3&"*",Wifi_LLAPI!$N:$N,"N/A")'
    )
    assert ws["J3"].value == "=IFERROR(E3/SUM(E3:F3),0)"
    assert ws["L3"].value == "=IFERROR(D3/C3,0)"
    assert ws["J3"].number_format == "0.00%"
    assert ws["L3"].number_format == "0.00%"

    assert ws["E9"].value == (
        '=COUNTIFS(Wifi_LLAPI!$A:$A,$B9&"*",Wifi_LLAPI!$O:$O,E$2)'
    )
    assert ws["E15"].value == (
        '=COUNTIFS(Wifi_LLAPI!$A:$A,$B15&"*",Wifi_LLAPI!$P:$P,E$2)'
    )
    assert ws["C8"].value == "=SUM(C3:C7)"
    assert ws["J8"].value == "=IFERROR(E8/SUM(E8:F8),0)"
    assert ws["L8"].value == "=IFERROR(D8/C8,0)"
    assert ws["O4"].value == "=L8"
    assert ws["P4"].value == "=IFERROR(SUM(E8:G8)/C8,0)"
    assert ws["O5"].value == "=L14"
    assert ws["P5"].value == "=IFERROR(SUM(E14:G14)/C14,0)"
    assert ws["O6"].value == "=L20"
    assert ws["P6"].value == "=IFERROR(SUM(E20:G20)/C20,0)"
    for cell in ("O4", "P4", "O5", "P5", "O6", "P6"):
        assert ws[cell].number_format == "0.00%"
    wb.close()

    wb = load_workbook(template, data_only=False)
    ws_wifi = wb["Wifi_LLAPI"]
    assert ws_wifi["N3"].value == "Summary Bucket WiFi 5G"
    assert ws_wifi["O3"].value == "Summary Bucket WiFi 6G"
    assert ws_wifi["P3"].value == "Summary Bucket WiFi 2.4G"
    assert ws_wifi.column_dimensions["N"].hidden is True
    assert ws_wifi.column_dimensions["O"].hidden is True
    assert ws_wifi.column_dimensions["P"].hidden is True
    wb.close()


# ---------------------------------------------------------------------------
# Task 2: template validation and summary sheet writer
# ---------------------------------------------------------------------------

def _create_template_with_summary(path: Path) -> None:
    """Create a minimal valid template workbook with Summary and Wifi_LLAPI sheets."""
    wb = Workbook()
    ws_summary = wb.active
    ws_summary.title = "Summary"
    ws_summary["A1"] = "Firmware Version"
    ws_summary["B1"] = "4.0.3"
    summary_headers = [
        "Module", "Object Category", "Total Items", "Tested Items",
        "Pass", "Fail", "To be confirmed", "Not Supported", "Skip",
        "Pass Rate", "result empty", "Progress",
    ]
    for col_idx, h in enumerate(summary_headers, start=1):
        ws_summary.cell(row=2, column=col_idx).value = h
    ws_summary["C1"] = "Detailed Statistics by Object"
    ws_summary.merge_cells("A3:A8")
    ws_summary["A3"] = "WiFi 5g"
    ws_summary["B3"] = "WiFi.AccessPoint"
    ws_summary["C3"] = '=COUNTIF(Wifi_LLAPI!$A:$A,$B3&"*")'
    ws_summary["D3"] = "=SUM(E3:I3)"
    ws_summary["E3"] = '=COUNTIFS(Wifi_LLAPI!$A:$A,$B3&"*",Wifi_LLAPI!$N:$N,E$2)'
    ws_summary["F3"] = '=COUNTIFS(Wifi_LLAPI!$A:$A,$B3&"*",Wifi_LLAPI!$N:$N,F$2)'
    ws_summary["G3"] = '=COUNTIFS(Wifi_LLAPI!$A:$A,$B3&"*",Wifi_LLAPI!$N:$N,G$2)'
    ws_summary["H3"] = (
        '=COUNTIFS(Wifi_LLAPI!$A:$A,$B3&"*",Wifi_LLAPI!$N:$N,H$2)'
        '+COUNTIFS(Wifi_LLAPI!$A:$A,$B3&"*",Wifi_LLAPI!$E:$E,'
        '"Not Supported",Wifi_LLAPI!$N:$N,"")'
    )
    ws_summary["I3"] = (
        '=COUNTIFS(Wifi_LLAPI!$A:$A,$B3&"*",Wifi_LLAPI!$N:$N,I$2)'
        '+COUNTIFS(Wifi_LLAPI!$A:$A,$B3&"*",Wifi_LLAPI!$N:$N,"N/A")'
    )
    ws_summary["J3"] = "=IFERROR(E3/SUM(E3:F3),0)"
    ws_summary["L3"] = "=IFERROR(D3/C3,0)"
    ws_summary["J3"].number_format = "0.00%"
    ws_summary["L3"].number_format = "0.00%"
    ws_summary["A9"] = "WiFi 6g"
    ws_summary["B9"] = "WiFi.AccessPoint"
    ws_summary["F9"] = '=COUNTIFS(Wifi_LLAPI!$A:$A,$B9&"*",Wifi_LLAPI!$O:$O,F$2)'
    ws_summary["J9"] = "=IFERROR(E9/SUM(E9:F9),0)"
    ws_summary["A15"] = "WiFi 2.4g"
    ws_summary["B15"] = "WiFi.AccessPoint"
    ws_summary["F15"] = '=COUNTIFS(Wifi_LLAPI!$A:$A,$B15&"*",Wifi_LLAPI!$P:$P,F$2)'
    ws_summary["J15"] = "=IFERROR(E15/SUM(E15:F15),0)"
    for row in range(3, 21):
        ws_summary[f"J{row}"] = f"=IFERROR(E{row}/SUM(E{row}:F{row}),0)"

    ws_wifi = wb.create_sheet("Wifi_LLAPI")
    row1_headers = [
        "Object", "Object Type", "Parameter Name", "HLAPI", "LLAPI",
        "Implemented by", "Test steps", "Command Output", None, None,
        None, "Tester", "Comment",
    ]
    for col_idx, h in enumerate(row1_headers, start=1):
        ws_wifi.cell(row=1, column=col_idx).value = h
    # row 2 is intentionally blank (None * 13)
    ws_wifi["I3"] = "WiFi 5G"
    ws_wifi["J3"] = "WiFi 6G"
    ws_wifi["K3"] = "WiFi 2.4G"
    ws_wifi["M3"] = "Comment"
    ws_wifi["N3"] = "Summary Bucket WiFi 5G"
    ws_wifi["O3"] = "Summary Bucket WiFi 6G"
    ws_wifi["P3"] = "Summary Bucket WiFi 2.4G"
    ws_wifi.column_dimensions["N"].hidden = True
    ws_wifi.column_dimensions["O"].hidden = True
    ws_wifi.column_dimensions["P"].hidden = True
    ws_wifi["A4"] = "WiFi.AccessPoint.{i}."
    ws_wifi["C4"] = "kickStation()"
    ws_wifi["A5"] = "WiFi.DataElements.Network.Device.{i}."
    ws_wifi["C5"] = "MaxNumMLDs"

    wb.save(path)
    wb.close()


def test_validate_wifi_llapi_template_accepts_valid(tmp_path: Path) -> None:
    p = tmp_path / "template.xlsx"
    _create_template_with_summary(p)
    result = validate_wifi_llapi_report_template(p)
    assert result == {"summary_sheet": "Summary", "wifi_sheet": "Wifi_LLAPI"}


def test_validate_wifi_llapi_template_rejects_missing_summary_bucket_columns(tmp_path: Path) -> None:
    p = tmp_path / "template.xlsx"
    _create_template_with_summary(p)
    wb = load_workbook(p)
    wb["Wifi_LLAPI"]["N3"] = None
    wb.save(p)
    wb.close()

    with pytest.raises(TemplateValidationError, match="Summary Bucket WiFi 5G"):
        validate_wifi_llapi_report_template(p)


def test_validate_wifi_llapi_template_rejects_raw_result_summary_fail_formula(tmp_path: Path) -> None:
    p = tmp_path / "template.xlsx"
    _create_template_with_summary(p)
    wb = load_workbook(p)
    wb["Summary"]["F3"] = '=COUNTIFS(Wifi_LLAPI!$A:$A,$B3&"*",Wifi_LLAPI!$I:$I,F$2)'
    wb.save(p)
    wb.close()

    with pytest.raises(TemplateValidationError, match=r"Summary.*F3.*N"):
        validate_wifi_llapi_report_template(p)


def test_validate_wifi_llapi_template_rejects_stale_pass_rate_denominator(tmp_path: Path) -> None:
    p = tmp_path / "template.xlsx"
    _create_template_with_summary(p)
    wb = load_workbook(p)
    wb["Summary"]["J8"] = "=IFERROR(E8/SUM(E8:G8),0)"
    wb.save(p)
    wb.close()

    with pytest.raises(TemplateValidationError, match=r"Summary.*J8.*E8:F8"):
        validate_wifi_llapi_report_template(p)


def test_validate_wifi_llapi_template_rejects_stale_summary_bucket_label(tmp_path: Path) -> None:
    p = tmp_path / "template.xlsx"
    _create_template_with_summary(p)
    wb = load_workbook(p)
    wb["Summary"]["B22"] = "Pass Rate is Pass / (Pass + Fail + To be tested)."
    wb.save(p)
    wb.close()

    with pytest.raises(TemplateValidationError, match=r"stale bucket label.*B22"):
        validate_wifi_llapi_report_template(p)


def test_validate_wifi_llapi_template_missing_summary_sheet(tmp_path: Path) -> None:
    p = tmp_path / "template.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Wifi_LLAPI"
    ws["I3"] = "WiFi 5G"
    ws["J3"] = "WiFi 6G"
    ws["K3"] = "WiFi 2.4G"
    wb.save(p)
    wb.close()
    with pytest.raises(TemplateValidationError, match="missing sheet: Summary"):
        validate_wifi_llapi_report_template(p)


def test_validate_wifi_llapi_template_result_header_mismatch(tmp_path: Path) -> None:
    p = tmp_path / "template.xlsx"
    _create_template_with_summary(p)
    wb = load_workbook(p)
    wb["Wifi_LLAPI"]["J3"] = "Wrong Value"
    wb.save(p)
    wb.close()
    with pytest.raises(TemplateValidationError, match=r"J.*WiFi 6G|WiFi 6G.*J"):
        validate_wifi_llapi_report_template(p)


def test_read_wifi_llapi_template_objects_carries_prefix(tmp_path: Path) -> None:
    p = tmp_path / "template.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Wifi_LLAPI"
    ws["A4"] = "WiFi.AccessPoint.{i}."
    ws["C4"] = "kickStation()"
    # A5 intentionally empty — carry-forward from A4 expected
    ws["C5"] = "MaxNumMLDs"
    wb.save(p)
    wb.close()

    objects = read_wifi_llapi_template_objects(p)
    assert objects[4] == "WiFi.AccessPoint.{i}."
    assert objects[5] == "WiFi.AccessPoint.{i}."


def test_write_summary_sheet_hybrid_layout(tmp_path: Path) -> None:
    p = tmp_path / "template.xlsx"
    _create_template_with_summary(p)

    summary_payload = {
        "policy_version": "wifi_llapi_summary_v1",
        "band_category": [
            {
                "band_key": "result_5g",
                "band_label": "5G",
                "category": "WiFi.AccessPoint",
                "total_items": 2,
                "tested_items": 2,
                "pass": 1,
                "fail": 0,
                "to_be_tested": 1,
                "not_supported": 0,
                "skip": 0,
                "pass_rate": 0.5,
                "progress": 1.0,
            }
        ],
    }

    result = write_summary_sheet(p, summary_payload)
    assert result == p

    wb = load_workbook(p, data_only=False)
    ws = wb["Summary"]
    assert ws["A1"].value == "Summary Policy"
    assert ws["B1"].value == "wifi_llapi_summary_v1"
    assert ws.cell(row=3, column=1).value == "Module"
    assert ws["A4"].value == "5G"
    assert ws["B4"].value == "WiFi.AccessPoint"
    assert ws["E4"].value == 1   # pass
    assert ws["G4"].value == 1   # to_be_tested
    assert ws["J4"].value == "=IFERROR(E4/SUM(E4:F4),0)"
    assert ws["L4"].value == "=IFERROR(D4/C4,0)"
    wb.close()


# ---------------------------------------------------------------------------
# Task 2 follow-up: spec-compliance fixes
# ---------------------------------------------------------------------------

def test_validate_wifi_llapi_template_column_a1_mismatch(tmp_path: Path) -> None:
    """A1 not containing 'Object' must raise TemplateValidationError."""
    p = tmp_path / "template.xlsx"
    _create_template_with_summary(p)
    wb = load_workbook(p)
    wb["Wifi_LLAPI"]["A1"] = "WrongHeader"
    wb.save(p)
    wb.close()
    with pytest.raises(TemplateValidationError, match=r"A.*Object|Object.*A"):
        validate_wifi_llapi_report_template(p)


def test_validate_wifi_llapi_template_missing_comment_column(tmp_path: Path) -> None:
    """M1 and M3 both lacking 'Comment' must raise TemplateValidationError."""
    p = tmp_path / "template.xlsx"
    _create_template_with_summary(p)
    wb = load_workbook(p)
    wb["Wifi_LLAPI"]["M1"] = None
    wb["Wifi_LLAPI"]["M3"] = None
    wb.save(p)
    wb.close()
    with pytest.raises(TemplateValidationError, match=r"[Cc]omment"):
        validate_wifi_llapi_report_template(p)


def test_write_summary_sheet_uses_payload_policy_version(tmp_path: Path) -> None:
    """payload['policy_version'] must be written to Summary!B1."""
    p = tmp_path / "template.xlsx"
    _create_template_with_summary(p)
    write_summary_sheet(p, {"policy_version": "custom_policy_v99", "band_category": []})
    wb = load_workbook(p)
    assert wb["Summary"]["B1"].value == "custom_policy_v99"
    wb.close()


def test_write_summary_sheet_fallback_policy_version(tmp_path: Path) -> None:
    """When payload lacks 'policy_version', SUMMARY_POLICY_VERSION is written."""
    from testpilot.reporting.wifi_llapi_summary import SUMMARY_POLICY_VERSION

    p = tmp_path / "template.xlsx"
    _create_template_with_summary(p)
    write_summary_sheet(p, {"band_category": []})
    wb = load_workbook(p)
    assert wb["Summary"]["B1"].value == SUMMARY_POLICY_VERSION
    wb.close()


# ---------------------------------------------------------------------------
# Task 2 quality fixes
# ---------------------------------------------------------------------------

def test_write_summary_sheet_closes_on_save_error(tmp_path: Path, monkeypatch) -> None:
    """wb.close() must be called even when wb.save() raises."""
    from unittest.mock import MagicMock
    from testpilot.reporting import wifi_llapi_excel

    fake_wb = MagicMock()
    fake_wb.sheetnames = []
    fake_wb.save.side_effect = OSError("disk full")

    monkeypatch.setattr(wifi_llapi_excel, "load_workbook", lambda *a, **kw: fake_wb)

    with pytest.raises(OSError, match="disk full"):
        write_summary_sheet(tmp_path / "x.xlsx", {"band_category": []})

    fake_wb.close.assert_called_once()


def test_write_summary_sheet_k_column_result_empty(tmp_path: Path) -> None:
    """Column K (result empty) must reflect unbucketed count, not hard-coded 0."""
    p = tmp_path / "template.xlsx"
    _create_template_with_summary(p)
    payload = {
        "band_category": [
            {
                "band_key": "result_5g",
                "band_label": "5G",
                "category": "WiFi.AccessPoint",
                "total_items": 2,
                "tested_items": 1,
                "pass": 1,
                "fail": 0,
                "to_be_tested": 0,
                "not_supported": 0,
                "skip": 0,
                "pass_rate": 1.0,
                "progress": 0.5,
            }
        ],
    }
    write_summary_sheet(p, payload)
    wb = load_workbook(p, data_only=False)
    # total_items=2, pass=1, rest=0 → 1 unbucketed; expect formula string
    assert wb["Summary"]["K4"].value == "=C4-SUM(E4:I4)"
    wb.close()


def test_validate_wifi_llapi_template_missing_wifi_sheet(tmp_path: Path) -> None:
    """Missing Wifi_LLAPI sheet must raise TemplateValidationError."""
    p = tmp_path / "template.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    summary_headers = [
        "Module", "Object Category", "Total Items", "Tested Items",
        "Pass", "Fail", "To be confirmed", "Not Supported", "Skip",
        "Pass Rate", "result empty", "Progress",
    ]
    for col_idx, h in enumerate(summary_headers, start=1):
        ws.cell(row=2, column=col_idx).value = h
    wb.save(p)
    wb.close()
    with pytest.raises(TemplateValidationError, match="missing sheet: Wifi_LLAPI"):
        validate_wifi_llapi_report_template(p)
