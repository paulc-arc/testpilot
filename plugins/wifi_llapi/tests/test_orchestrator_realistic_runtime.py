"""Realistic integration tests for orchestrator wifi_llapi runtime path."""

from __future__ import annotations

import copy
import json
from pathlib import Path
import shutil
from typing import Any

from openpyxl import Workbook, load_workbook
import pytest

from testpilot.core.orchestrator import Orchestrator

FAIL_CASE_ID = "wifi-llapi-D004-evaluate-fail"
PASS_CASE_ID = "wifi-llapi-D005-evaluate-pass"
PASS_TOKEN = "TOKEN_OK"


class MockTransport:
    """Test-only transport mock to simulate command execution."""

    def __init__(self) -> None:
        self._connected = False
        self.history: list[str] = []

    @property
    def is_connected(self) -> bool:
        return self._connected

    def connect(self) -> None:
        self._connected = True

    def disconnect(self) -> None:
        self._connected = False

    def execute(self, command: str, timeout: float = 30.0) -> dict[str, Any]:
        self.history.append(command)
        return {
            "returncode": 0,
            "stdout": f"[mock] timeout={timeout} {command}",
            "stderr": "",
            "elapsed": 0.01,
        }


def _write_testbed_yaml(path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(
        """
testbed:
  name: integration-testbed
  serialwrap_binary: /tmp/serialwrap
  devices:
    DUT:
      role: ap
      transport: serial
    STA:
      role: sta
      transport: adb
""".strip()
        + "\n",
        encoding="utf-8",
    )


def _write_source_xlsx(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Wifi_LLAPI"

    ws["A2"] = "Object"
    ws["C2"] = "Parameter Name"
    ws["G2"] = "Test steps"
    ws["H2"] = "Command Output"
    ws["I2"] = "ARC 4.0.3 Test Result\nWiFi 5g"
    ws["J2"] = "ARC 4.0.3 Test Result\nWiFi 6g"
    ws["K2"] = "ARC 4.0.3 Test Result\nWiFi 2.4g"
    ws["L2"] = "Tester"
    ws["M2"] = "ARC Comment"

    ws["A4"] = "WiFi.AccessPoint.{i}."
    ws["C4"] = "kickStation()"
    ws["A5"] = "WiFi.Radio.{i}."
    ws["C5"] = "getRadioStats()"
    ws["A6"] = "WiFi.AccessPoint.{i}.AssociatedDevice.{i}."
    ws["C6"] = "Noise"

    wb.save(path)
    wb.close()


def _prepare_runtime_project(tmp_path: Path) -> tuple[Path, Path]:
    project_root = tmp_path / "project"
    plugin_dir = project_root / "plugins" / "wifi_llapi"
    plugin_dir.mkdir(parents=True, exist_ok=True)

    repo_root = Path(__file__).resolve().parents[3]
    source_plugin_dir = repo_root / "plugins" / "wifi_llapi"
    shutil.copy2(source_plugin_dir / "plugin.py", plugin_dir / "plugin.py")
    shutil.copy2(source_plugin_dir / "command_resolver.py", plugin_dir / "command_resolver.py")
    shutil.copy2(source_plugin_dir / "agent-config.yaml", plugin_dir / "agent-config.yaml")
    shutil.copy2(source_plugin_dir / "band-baselines.yaml", plugin_dir / "band-baselines.yaml")
    shutil.copy2(source_plugin_dir / "baseline_qualifier.py", plugin_dir / "baseline_qualifier.py")

    _write_testbed_yaml(project_root / "configs" / "testbed.yaml")
    source_xlsx = project_root / "source.xlsx"
    _write_source_xlsx(source_xlsx)
    return project_root, source_xlsx


def _build_cases() -> list[dict[str, Any]]:
    return [
        {
            "id": FAIL_CASE_ID,
            "name": "evaluate-fail",
            "source": {
                "row": 4,
                "object": "WiFi.AccessPoint.{i}.",
                "api": "kickStation()",
            },
            "steps": [
                {
                    "id": "step1",
                    "target": "DUT",
                    "command": 'ubus-cli "WiFi.AccessPoint.1.kickStation(macaddress=AA:BB:CC:DD:EE:FF)"',
                }
            ],
            "pass_criteria": [
                {
                    "field": "step1.output",
                    "operator": "contains",
                    "value": PASS_TOKEN,
                }
            ],
            "emit_pass_token": False,
            "bands": ["5g"],
        },
        {
            "id": PASS_CASE_ID,
            "name": "evaluate-pass",
            "source": {
                "row": 5,
                "object": "WiFi.Radio.{i}.",
                "api": "getRadioStats()",
            },
            "steps": [
                {
                    "id": "step1",
                    "target": "STA",
                    "command": 'ubus-cli "WiFi.Radio.1.getRadioStats()"',
                }
            ],
            "pass_criteria": [
                {
                    "field": "step1.output",
                    "operator": "contains",
                    "value": PASS_TOKEN,
                }
            ],
            "emit_pass_token": True,
            "bands": ["6g"],
        },
    ]


def _patch_runtime_hooks(monkeypatch, plugin: Any, cases: list[dict[str, Any]]) -> dict[str, Any]:
    transports = {
        "DUT": MockTransport(),
        "STA": MockTransport(),
    }
    state: dict[str, Any] = {
        "setup_calls": [],
        "verify_calls": [],
        "execute_calls": [],
        "evaluate_calls": [],
        "teardown_calls": [],
        "transports": transports,
    }

    def discover_cases() -> list[dict[str, Any]]:
        return copy.deepcopy(cases)

    def setup_env(case: dict[str, Any], topology: Any) -> bool:
        state["setup_calls"].append(str(case.get("id", "?")))
        assert "DUT" in topology.devices
        assert "STA" in topology.devices
        for transport in transports.values():
            if not transport.is_connected:
                transport.connect()
        return True

    def verify_env(case: dict[str, Any], topology: Any) -> bool:
        case_id = str(case.get("id", "?"))
        state["verify_calls"].append(case_id)
        probe = transports["DUT"].execute(f"verify:{case_id}", timeout=5.0)
        return probe["returncode"] == 0

    def execute_step(case: dict[str, Any], step: dict[str, Any], topology: Any) -> dict[str, Any]:
        case_id = str(case.get("id", "?"))
        step_id = str(step.get("id", "step"))
        attempt = int(case.get("_attempt_index", 0))
        target = str(step.get("target", "DUT"))
        timeout = float(step.get("timeout", 0.0))
        command = str(step.get("command", ""))
        transport = transports[target]

        result = transport.execute(command, timeout=timeout)
        output = str(result.get("stdout", ""))
        if bool(case.get("emit_pass_token")):
            output = (
                f"root@prplOS:/# {command}\n"
                f"{output} {PASS_TOKEN}\n"
                ">\n"
            )
        else:
            output = (
                f"root@prplOS:/# {command}\n"
                f"{output} TOKEN_MISSING\n"
                ">\n"
            )

        state["execute_calls"].append(
            {
                "case_id": case_id,
                "attempt": attempt,
                "step_id": step_id,
                "target": target,
                "timeout": timeout,
            }
        )
        return {
            "success": True,
            "output": output,
            "captured": {"stdout": result.get("stdout", "")},
            "timing": float(result.get("elapsed", 0.0)),
        }

    def evaluate(case: dict[str, Any], results: dict[str, Any]) -> bool:
        case_id = str(case.get("id", "?"))
        attempt = int(case.get("_attempt_index", 0))
        state["evaluate_calls"].append({"case_id": case_id, "attempt": attempt})

        steps = results.get("steps", {})
        if not isinstance(steps, dict):
            return False

        for criterion in case.get("pass_criteria", []):
            if not isinstance(criterion, dict):
                continue
            field = str(criterion.get("field", ""))
            operator = str(criterion.get("operator", ""))
            expected = str(criterion.get("value", ""))

            step_key, _, attr = field.partition(".")
            step_result = steps.get(step_key, {})
            if not isinstance(step_result, dict):
                return False
            actual = str(step_result.get(attr, ""))
            if operator == "contains" and expected not in actual:
                return False
        return True

    def teardown(case: dict[str, Any], topology: Any) -> None:
        state["teardown_calls"].append(str(case.get("id", "?")))
        for transport in transports.values():
            if transport.is_connected:
                transport.disconnect()

    monkeypatch.setattr(plugin, "discover_cases", discover_cases)
    monkeypatch.setattr(plugin, "setup_env", setup_env)
    monkeypatch.setattr(plugin, "verify_env", verify_env)
    monkeypatch.setattr(plugin, "execute_step", execute_step)
    monkeypatch.setattr(plugin, "evaluate", evaluate)
    monkeypatch.setattr(plugin, "teardown", teardown)
    return state


def _load_case_traces(trace_dir: Path) -> dict[str, dict[str, Any]]:
    traces: dict[str, dict[str, Any]] = {}
    for path in sorted(trace_dir.glob("*.json")):
        payload = json.loads(path.read_text(encoding="utf-8"))
        case_id = payload.get("case_id")
        if isinstance(case_id, str):
            traces[case_id] = payload
    return traces


def _report_snapshot(report_path: Path) -> dict[int, dict[str, Any]]:
    wb = load_workbook(report_path)
    ws = wb["Wifi_LLAPI"]
    snapshot = {
        4: {column: ws[f"{column}4"].value for column in ("G", "H", "I", "J", "K", "L")},
        5: {column: ws[f"{column}5"].value for column in ("G", "H", "I", "J", "K", "L")},
    }
    wb.close()
    return snapshot


def _normalize_trace_payload(payload: dict[str, Any]) -> dict[str, Any]:
    selection_trace = payload.get("selection_trace", {})
    attempts = payload.get("attempts", [])
    session_plan = selection_trace.get("session_plan", {})
    normalized_attempts = [
        {
            "attempt": item.get("attempt"),
            "timeout_seconds": item.get("timeout_seconds"),
            "runner": item.get("runner"),
            "status": item.get("status"),
            "comment": item.get("comment"),
            "commands": item.get("commands"),
            "outputs": item.get("outputs"),
        }
        for item in attempts
    ]
    return {
        "case_id": payload.get("case_id"),
        "source_row": payload.get("source_row"),
        "execution": payload.get("execution"),
        "selected": selection_trace.get("selected"),
        "fallback": selection_trace.get("fallback"),
        "runtime": {
            "method": selection_trace.get("runtime", {}).get("method"),
            "reason": selection_trace.get("runtime", {}).get("reason"),
            "selection": selection_trace.get("runtime", {}).get("selection"),
        },
        "session_plan": {
            "provider": session_plan.get("provider"),
            "model": session_plan.get("model"),
            "reasoning_effort": session_plan.get("reasoning_effort"),
            "status": session_plan.get("status"),
        },
        "attempts": normalized_attempts,
        "final": payload.get("final"),
    }


def _runtime_snapshot(result: dict[str, Any]) -> dict[str, Any]:
    trace_dir = Path(result["agent_trace_dir"])
    traces = _load_case_traces(trace_dir)
    return {
        "cases_count": result["cases_count"],
        "pass_count": result["pass_count"],
        "fail_count": result["fail_count"],
        "report": _report_snapshot(Path(result["report_path"])),
        "traces": {
            case_id: _normalize_trace_payload(payload)
            for case_id, payload in sorted(traces.items())
        },
    }


def _run_realistic_runtime(tmp_path: Path, monkeypatch) -> tuple[dict[str, Any], dict[str, Any]]:
    project_root, source_xlsx = _prepare_runtime_project(tmp_path)
    orch = Orchestrator(
        project_root=project_root,
        plugins_dir=project_root / "plugins",
        config_path=project_root / "configs" / "testbed.yaml",
    )

    plugin = orch.loader.load("wifi_llapi")
    cases = _build_cases()
    state = _patch_runtime_hooks(monkeypatch, plugin=plugin, cases=cases)

    result = orch.run(
        "wifi_llapi",
        case_ids=[FAIL_CASE_ID, PASS_CASE_ID],
        dut_fw_ver="FW-IT-REALISTIC-1",
        report_source_xlsx=str(source_xlsx),
    )
    return result, state


def test_realistic_runtime_uses_results_reference_for_band_specific_statuses(
    tmp_path: Path,
    monkeypatch,
):
    project_root, source_xlsx = _prepare_runtime_project(tmp_path)
    orch = Orchestrator(
        project_root=project_root,
        plugins_dir=project_root / "plugins",
        config_path=project_root / "configs" / "testbed.yaml",
    )
    plugin = orch.loader.load("wifi_llapi")
    cases = [
        {
            "id": "wifi-llapi-D029-mode-reference-pass",
            "name": "mixed-status-pass",
            "source": {
                "row": 4,
                "object": "WiFi.AccessPoint.{i}.",
                "api": "kickStation()",
                "baseline": "BCM v4.0.3",
            },
            "steps": [
                {
                    "id": "step1",
                    "target": "DUT",
                    "command": 'ubus-cli "WiFi.AccessPoint.1.kickStation(macaddress=AA:BB:CC:DD:EE:FF)"',
                }
            ],
            "pass_criteria": [
                {
                    "field": "step1.output",
                    "operator": "contains",
                    "value": PASS_TOKEN,
                }
            ],
            "emit_pass_token": True,
            "results_reference": {
                "v4.0.3": {
                    "5g": "Not Supported",
                    "6g": "Skip",
                    "2.4g": "Not Supported",
                }
            },
        },
        {
            "id": "wifi-llapi-D023-inactive-reference-fail",
            "name": "mixed-status-fail",
            "source": {
                "row": 5,
                "object": "WiFi.Radio.{i}.",
                "api": "getRadioStats()",
                "baseline": "BCM v4.0.3",
            },
            "steps": [
                {
                    "id": "step1",
                    "target": "DUT",
                    "command": 'ubus-cli "WiFi.Radio.1.getRadioStats()"',
                }
            ],
            "pass_criteria": [
                {
                    "field": "step1.output",
                    "operator": "contains",
                    "value": PASS_TOKEN,
                }
            ],
            "emit_pass_token": False,
            "results_reference": {
                "v4.0.3": {
                    "5g": "Pass",
                    "6g": "Skip",
                    "2.4g": "Pass",
                }
            },
        },
        {
            "id": "wifi-llapi-D034-noise-reference-single-band",
            "name": "single-band-fail",
            "bands": ["5g"],
            "source": {
                "row": 6,
                "object": "WiFi.AccessPoint.{i}.AssociatedDevice.{i}.",
                "api": "Noise",
                "baseline": "BCM v4.0.3",
            },
            "steps": [
                {
                    "id": "step1",
                    "target": "DUT",
                    "command": 'ubus-cli "WiFi.AccessPoint.1.AssociatedDevice.1.Noise?"',
                }
            ],
            "pass_criteria": [
                {
                    "field": "step1.output",
                    "operator": "contains",
                    "value": PASS_TOKEN,
                }
            ],
            "emit_pass_token": True,
            "results_reference": {
                "v4.0.3": {
                    "5g": "Fail",
                    "6g": "N/A",
                    "2.4g": "N/A",
                }
            },
        },
    ]
    _patch_runtime_hooks(monkeypatch, plugin=plugin, cases=cases)

    result = orch.run(
        "wifi_llapi",
        case_ids=[case["id"] for case in cases],
        dut_fw_ver="FW-IT-REALISTIC-REF-1",
        report_source_xlsx=str(source_xlsx),
    )

    assert result["status"] == "completed"
    assert result["pass_count"] == 1
    assert result["fail_count"] == 2

    wb = load_workbook(Path(result["report_path"]))
    ws = wb["Wifi_LLAPI"]
    assert ws["I4"].value == "Not Supported"
    assert ws["J4"].value == "Skip"
    assert ws["K4"].value == "Not Supported"
    assert ws["I5"].value == "Fail"
    assert ws["J5"].value == "Skip"
    assert ws["K5"].value == "Fail"
    assert ws["I6"].value == "Fail"
    assert ws["J6"].value == "N/A"
    assert ws["K6"].value == "N/A"
    wb.close()

    traces = _load_case_traces(Path(result["agent_trace_dir"]))
    single_band_trace = traces["wifi-llapi-D034-noise-reference-single-band"]
    assert single_band_trace["final"]["status"] == "Fail"
    assert single_band_trace["final"]["evaluation_verdict"] == "Pass"
    assert single_band_trace["attempts"][0]["status"] == "Fail"
    assert single_band_trace["attempts"][0]["evaluation_verdict"] == "Pass"


def test_realistic_runtime_covers_hooks_and_report_outputs(tmp_path: Path, monkeypatch):
    result, state = _run_realistic_runtime(tmp_path, monkeypatch)

    assert result["status"] == "completed"
    assert result["cases_count"] == 2
    assert result["pass_count"] == 1
    assert result["fail_count"] == 1
    assert result["agent_trace_count"] == 2

    report_path = Path(result["report_path"])
    md_report_path = Path(result["md_report_path"])
    json_report_path = Path(result["json_report_path"])
    trace_dir = Path(result["agent_trace_dir"])
    artifact_dir = Path(result["artifact_dir"])
    assert report_path.is_file()
    assert md_report_path.is_file()
    assert json_report_path.is_file()
    assert trace_dir.is_dir()
    assert artifact_dir.is_dir()
    assert result["run_id"] in report_path.name
    assert artifact_dir == report_path.parent
    assert md_report_path.parent == artifact_dir
    assert json_report_path.parent == artifact_dir
    assert trace_dir.parent == artifact_dir
    assert artifact_dir.name == report_path.stem

    # 1 fail case (2 attempts) + 1 pass case (1 attempt) = 3 executions per hook.
    assert len(state["setup_calls"]) == 3
    assert len(state["verify_calls"]) == 3
    assert len(state["execute_calls"]) == 3
    assert len(state["evaluate_calls"]) == 3
    assert len(state["teardown_calls"]) == 3

    wb = load_workbook(report_path)
    ws = wb["Wifi_LLAPI"]
    assert ws["G4"].value == 'ubus-cli "WiFi.AccessPoint.1.kickStation(macaddress=AA:BB:CC:DD:EE:FF)"'
    assert "root@prplOS" not in str(ws["H4"].value)
    assert ws["I4"].value == "Fail"
    assert ws["J4"].value == "N/A"
    assert ws["K4"].value == "N/A"
    assert ws["L4"].value == "testpilot"
    assert ws.max_column == 12

    assert ws["G5"].value == 'ubus-cli "WiFi.Radio.1.getRadioStats()"'
    assert PASS_TOKEN in str(ws["H5"].value)
    assert "root@prplOS" not in str(ws["H5"].value)
    assert ws["I5"].value == "N/A"
    assert ws["J5"].value == "Pass"
    assert ws["K5"].value == "N/A"
    assert ws["L5"].value == "testpilot"

    meta = wb["_meta"]
    assert meta["B2"].value == "FW-IT-REALISTIC-1"
    wb.close()

    report_text = md_report_path.read_text(encoding="utf-8")
    assert "## Timing" in report_text
    assert "## Suite summary" in report_text
    assert "## Per-case timing" in report_text
    assert "| pass_cases | failed_cases | other_cases | pass_rate |" in report_text
    assert "| 0 | 1 | 1 | `0.00%` |" in report_text
    assert "| environment buildup |" in report_text
    assert f"| {FAIL_CASE_ID} |" in report_text
    assert f"| {PASS_CASE_ID} |" in report_text

    transports = state["transports"]
    assert any(cmd.startswith("verify:") for cmd in transports["DUT"].history)
    assert any("kickStation" in cmd for cmd in transports["DUT"].history)
    assert any("getRadioStats" in cmd for cmd in transports["STA"].history)


def test_fail_and_continue_with_plugin_evaluate_failure(tmp_path: Path, monkeypatch):
    result, state = _run_realistic_runtime(tmp_path, monkeypatch)

    assert result["status"] == "completed"
    assert result["pass_count"] == 1
    assert result["fail_count"] == 1

    execute_case_order = [item["case_id"] for item in state["execute_calls"]]
    assert execute_case_order == [FAIL_CASE_ID, FAIL_CASE_ID, PASS_CASE_ID]

    evaluate_case_order = [item["case_id"] for item in state["evaluate_calls"]]
    assert evaluate_case_order == [FAIL_CASE_ID, FAIL_CASE_ID, PASS_CASE_ID]

    traces = _load_case_traces(Path(result["agent_trace_dir"]))
    assert set(traces.keys()) == {FAIL_CASE_ID, PASS_CASE_ID}

    fail_trace = traces[FAIL_CASE_ID]
    pass_trace = traces[PASS_CASE_ID]

    assert fail_trace["final"]["status"] == "Fail"
    assert fail_trace["final"]["evaluation_verdict"] == "Fail"
    assert fail_trace["final"]["attempts_used"] == 2
    assert len(fail_trace["attempts"]) == 2
    assert all(item["status"] == "Fail" for item in fail_trace["attempts"])
    assert all(item["evaluation_verdict"] == "Fail" for item in fail_trace["attempts"])
    assert "pass_criteria not satisfied" in str(fail_trace["final"]["comment"])

    assert pass_trace["final"]["status"] == "Pass"
    assert pass_trace["final"]["evaluation_verdict"] == "Pass"
    assert pass_trace["final"]["attempts_used"] == 1
    assert len(pass_trace["attempts"]) == 1
    assert pass_trace["attempts"][0]["status"] == "Pass"
    assert pass_trace["attempts"][0]["evaluation_verdict"] == "Pass"
    assert pass_trace["selection_trace"]["session_plan"]["model"] == "gpt-5.4"
    assert pass_trace["selection_trace"]["session_plan"]["status"] == "planned"


def test_realistic_runtime_report_paths_remain_unique_across_runs(tmp_path: Path, monkeypatch):
    project_root, source_xlsx = _prepare_runtime_project(tmp_path)
    orch = Orchestrator(
        project_root=project_root,
        plugins_dir=project_root / "plugins",
        config_path=project_root / "configs" / "testbed.yaml",
    )
    plugin = orch.loader.load("wifi_llapi")
    cases = _build_cases()
    _patch_runtime_hooks(monkeypatch, plugin=plugin, cases=cases)

    first_result = orch.run(
        "wifi_llapi",
        case_ids=[FAIL_CASE_ID, PASS_CASE_ID],
        dut_fw_ver="FW-IT-REALISTIC-1",
        report_source_xlsx=str(source_xlsx),
    )
    second_result = orch.run(
        "wifi_llapi",
        case_ids=[FAIL_CASE_ID, PASS_CASE_ID],
        dut_fw_ver="FW-IT-REALISTIC-1",
        report_source_xlsx=str(source_xlsx),
    )

    assert first_result["report_path"] != second_result["report_path"]
    assert first_result["artifact_dir"] != second_result["artifact_dir"]


def test_realistic_runtime_results_remain_identical_across_three_runs(tmp_path: Path, monkeypatch):
    project_root, source_xlsx = _prepare_runtime_project(tmp_path)
    orch = Orchestrator(
        project_root=project_root,
        plugins_dir=project_root / "plugins",
        config_path=project_root / "configs" / "testbed.yaml",
    )
    plugin = orch.loader.load("wifi_llapi")
    cases = _build_cases()
    _patch_runtime_hooks(monkeypatch, plugin=plugin, cases=cases)

    first_result = orch.run(
        "wifi_llapi",
        case_ids=[FAIL_CASE_ID, PASS_CASE_ID],
        dut_fw_ver="FW-IT-REALISTIC-1",
        report_source_xlsx=str(source_xlsx),
    )
    second_result = orch.run(
        "wifi_llapi",
        case_ids=[FAIL_CASE_ID, PASS_CASE_ID],
        dut_fw_ver="FW-IT-REALISTIC-1",
        report_source_xlsx=str(source_xlsx),
    )
    third_result = orch.run(
        "wifi_llapi",
        case_ids=[FAIL_CASE_ID, PASS_CASE_ID],
        dut_fw_ver="FW-IT-REALISTIC-1",
        report_source_xlsx=str(source_xlsx),
    )

    first_snapshot = _runtime_snapshot(first_result)
    second_snapshot = _runtime_snapshot(second_result)
    third_snapshot = _runtime_snapshot(third_result)

    assert first_snapshot == second_snapshot == third_snapshot


def test_realistic_runtime_can_rerun_from_existing_template_without_source_xlsx(
    tmp_path: Path,
    monkeypatch,
):
    project_root, source_xlsx = _prepare_runtime_project(tmp_path)
    orch = Orchestrator(
        project_root=project_root,
        plugins_dir=project_root / "plugins",
        config_path=project_root / "configs" / "testbed.yaml",
    )
    plugin = orch.loader.load("wifi_llapi")
    cases = _build_cases()
    _patch_runtime_hooks(monkeypatch, plugin=plugin, cases=cases)

    first_result = orch.run(
        "wifi_llapi",
        case_ids=[FAIL_CASE_ID, PASS_CASE_ID],
        dut_fw_ver="FW-IT-REALISTIC-1",
        report_source_xlsx=str(source_xlsx),
    )
    Path(source_xlsx).unlink()

    second_result = orch.run(
        "wifi_llapi",
        case_ids=[FAIL_CASE_ID, PASS_CASE_ID],
        dut_fw_ver="FW-IT-REALISTIC-1",
    )

    assert first_result["status"] == "completed"
    assert second_result["status"] == "completed"
    assert Path(second_result["template_path"]).is_file()
    assert Path(second_result["report_path"]).is_file()
    # Manifest now stores portable relative paths, so compare basenames
    assert Path(second_result["source_report"]).name == Path(str(source_xlsx)).name


def test_realistic_runtime_missing_source_xlsx_does_not_create_artifact_dir(
    tmp_path: Path,
    monkeypatch,
):
    project_root, _source_xlsx = _prepare_runtime_project(tmp_path)
    orch = Orchestrator(
        project_root=project_root,
        plugins_dir=project_root / "plugins",
        config_path=project_root / "configs" / "testbed.yaml",
    )
    plugin = orch.loader.load("wifi_llapi")
    cases = _build_cases()
    _patch_runtime_hooks(monkeypatch, plugin=plugin, cases=cases)

    missing_source_xlsx = project_root / "missing.xlsx"
    with pytest.raises(FileNotFoundError, match="wifi_llapi source report not found"):
        orch.run(
            "wifi_llapi",
            case_ids=[FAIL_CASE_ID, PASS_CASE_ID],
            dut_fw_ver="FW-IT-REALISTIC-1",
            report_source_xlsx=str(missing_source_xlsx),
        )

    assert not (project_root / "plugins" / "wifi_llapi" / "reports").exists()


def test_realistic_runtime_records_pass_after_remediation(tmp_path: Path, monkeypatch):
    project_root, source_xlsx = _prepare_runtime_project(tmp_path)
    orch = Orchestrator(
        project_root=project_root,
        plugins_dir=project_root / "plugins",
        config_path=project_root / "configs" / "testbed.yaml",
    )
    plugin = orch.loader.load("wifi_llapi")
    cases = [
        {
            "id": "wifi-llapi-D900-remediation-pass",
            "name": "remediation-pass",
            "source": {
                "row": 4,
                "object": "WiFi.AccessPoint.{i}.",
                "api": "kickStation()",
            },
            "steps": [
                {
                    "id": "step1",
                    "target": "DUT",
                    "command": 'ubus-cli "WiFi.AccessPoint.1.kickStation(macaddress=AA:BB:CC:DD:EE:FF)"',
                }
            ],
            "pass_criteria": [
                {
                    "field": "step1.output",
                    "operator": "contains",
                    "value": PASS_TOKEN,
                }
            ],
            "emit_pass_token": True,
            "bands": ["5g"],
        }
    ]
    state = _patch_runtime_hooks(monkeypatch, plugin=plugin, cases=cases)
    verify_attempts = {"count": 0}
    remediation_calls: list[dict[str, Any]] = []

    def verify_env(case: dict[str, Any], topology: Any) -> bool:
        del topology
        verify_attempts["count"] += 1
        if verify_attempts["count"] == 1:
            case["_last_failure"] = {
                "case_id": case["id"],
                "attempt_index": int(case.get("_attempt_index", 1)),
                "phase": "verify_env",
                "comment": "STA band baseline/connect failed",
                "category": "environment",
                "reason_code": "sta_band_not_ready",
                "band": "5g",
                "device": "STA",
            }
            return False
        return True

    def build_remediation_decision(case: dict[str, Any], failure_snapshot: Any, topology: Any, *, runner=None, remediation_policy=None):
        del topology, runner, remediation_policy
        return {
            "case_id": case["id"],
            "attempt_index": failure_snapshot.attempt_index,
            "summary": "repair baseline",
            "actions": [
                {"executor_key": "sta_band_rebaseline", "band": "5g"},
                {"executor_key": "case_env_reverify"},
            ],
        }

    def execute_remediation(case: dict[str, Any], decision: Any, topology: Any) -> dict[str, Any]:
        del case, topology
        remediation_calls.append(decision.to_dict())
        return {
            "success": True,
            "verify_after": True,
            "comment": "baseline repaired",
            "actions": [
                {"executor_key": "sta_band_rebaseline", "success": True},
                {"executor_key": "case_env_reverify", "success": True},
            ],
        }

    monkeypatch.setattr(plugin, "verify_env", verify_env)
    monkeypatch.setattr(plugin, "build_remediation_decision", build_remediation_decision)
    monkeypatch.setattr(plugin, "execute_remediation", execute_remediation)

    result = orch.run(
        "wifi_llapi",
        case_ids=[cases[0]["id"]],
        dut_fw_ver="FW-IT-REMEDIATION-1",
        report_source_xlsx=str(source_xlsx),
    )

    assert result["status"] == "completed"
    assert result["pass_count"] == 1
    assert verify_attempts["count"] == 2
    assert len(remediation_calls) == 1

    traces = _load_case_traces(Path(result["agent_trace_dir"]))
    trace = traces["wifi-llapi-D900-remediation-pass"]
    assert trace["final"]["diagnostic_status"] == "PassAfterRemediation"
    assert trace["final"]["attempts_used"] == 2
    assert len(trace["remediation_history"]) == 1
    assert trace["remediation_history"][0]["applied"] is True
    assert trace["remediation_history"][0]["verify_after"] is True
    assert state["evaluate_calls"][-1]["case_id"] == "wifi-llapi-D900-remediation-pass"
