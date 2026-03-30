#!/usr/bin/env python3
"""
Generate YAML case descriptors from Wifi_LLAPI xlsx sheet.
v4: Proper disambiguation with readable filenames.
    Strategy: build human-readable suffix FIRST, then add D{row} prefix.
    Collision resolution: hlapi_value → deeper object context → guaranteed unique.

Usage:
    python scripts/gen_cases.py --xlsx <path-to-xlsx> [--output-dir <dir>]
"""
import argparse
import re, yaml, sys
from pathlib import Path
from collections import Counter, defaultdict

from openpyxl import load_workbook


def _project_root() -> Path:
    return Path(__file__).resolve().parents[1]


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Generate YAML case descriptors from Wifi_LLAPI xlsx sheet.",
    )
    parser.add_argument(
        "--xlsx", required=True, type=Path,
        help="Path to the source LLAPI Test Report xlsx file.",
    )
    parser.add_argument(
        "--output-dir", type=Path, default=None,
        help="Output directory for generated YAML files "
             "(default: plugins/wifi_llapi/cases/ relative to project root).",
    )
    args = parser.parse_args()

    xlsx_path: Path = args.xlsx
    if not xlsx_path.exists():
        print(f"ERROR: xlsx file not found: {xlsx_path}", file=sys.stderr)
        raise SystemExit(1)

    out_dir: Path = args.output_dir or (_project_root() / "plugins" / "wifi_llapi" / "cases")
    out_dir.mkdir(parents=True, exist_ok=True)

    _generate(xlsx_path, out_dir)


def _generate(xlsx_path: Path, out_dir: Path) -> None:
    wb = load_workbook(xlsx_path, data_only=True, read_only=True)
    ws = wb['Wifi_LLAPI']
    rows_data = list(ws.iter_rows(min_row=1, values_only=False))

    def cell_str(cell):
        v = cell.value
        return str(v).strip() if v is not None else ''

    def sanitize(s, max_len=50):
        s = re.sub(r'[^a-z0-9]+', '_', s.lower()).strip('_')
        return s[:max_len].rstrip('_')

    def parse_steps(raw):
        if not raw: return []
        lines = [l.strip() for l in raw.split('\n') if l.strip()]
        steps, buf = [], []
        for line in lines:
            if re.match(r'^(\d+[\.\)]\s|##\s*\d|\\\\)', line) and buf:
                steps.append(' '.join(buf))
                buf = [re.sub(r'^(\d+[\.\)]\s*|##\s*\d+[\.\)]*\s*|\\\\+\s*)', '', line)]
            else:
                buf.append(re.sub(r'^(root@\w+:/?[#$]\s*)', '', line))
        if buf: steps.append(' '.join(buf))
        return steps

    def detect_topology(text):
        sta = bool(re.search(r'\b(STA|station|client|connect.*wifi|arping|wl\d)', text, re.I))
        ep = bool(re.search(r'\b(endpoint|iperf|server|PC|host)', text, re.I))
        return sta, ep

    def result_str(c):
        v = cell_str(c)
        return '' if (not v or v == 'None') else v

    # ── Collect entries ──
    entries = []
    current_obj = ''
    for row in rows_data[1:]:
        obj = cell_str(row[0])
        param = cell_str(row[1])
        if not param: continue
        if obj: current_obj = obj
        rn = row[0].row
        hlapi = cell_str(row[2])
        entries.append({
            'row': rn, 'obj': current_obj, 'param': param,
            'hlapi': hlapi,
            'llapi_sup': cell_str(row[3]), 'impl_by': cell_str(row[4]),
            'steps_raw': cell_str(row[5]), 'verify_cmd': cell_str(row[6]),
            'r401': {'5g': result_str(row[7]) if len(row)>7 else '',
                     '6g': result_str(row[8]) if len(row)>8 else '',
                     '2.4g': result_str(row[9]) if len(row)>9 else '',
                     'comment': result_str(row[10]) if len(row)>10 else ''},
            'r403': {'5g': result_str(row[11]) if len(row)>11 else '',
                     '6g': result_str(row[12]) if len(row)>12 else '',
                     '2.4g': result_str(row[13]) if len(row)>13 else '',
                     'comment': result_str(row[14]) if len(row)>14 else ''},
        })

    # ── Build readable suffix (without row prefix) ──
    def build_suffix(e):
        """Build a descriptive suffix for this entry."""
        base = sanitize(re.sub(r'[()]', '', e['param']))
        hlapi = e['hlapi']
        obj = e['obj']

        obj_parts = [p for p in obj.rstrip('.').split('.') if p and '{' not in p and p.lower() != 'wifi']

        m = re.search(r'\(\)\s*(.+?)(?:\s*[:,=]|$)', hlapi)
        subfield = sanitize(m.group(1).strip()) if m and m.group(1).strip() else None

        m = re.search(r'=\s*(\S+)', hlapi)
        hlapi_val = sanitize(m.group(1).rstrip(',')) if m else None

        return base, obj_parts, subfield, hlapi_val

    # First pass: count how many entries share each base param name
    base_counter = Counter()
    for e in entries:
        base = sanitize(re.sub(r'[()]', '', e['param']))
        base_counter[base] += 1

    # Second pass: build suffix for each entry
    suffixes = {}
    for e in entries:
        base, obj_parts, subfield, hlapi_val = build_suffix(e)

        if base_counter[base] == 1:
            suffixes[e['row']] = base
        elif subfield and subfield != base:
            suffixes[e['row']] = f"{base}_{subfield}"
        elif obj_parts:
            ctx = [sanitize(p) for p in obj_parts if sanitize(p) != base]
            if ctx:
                suffixes[e['row']] = f"{base}_{'_'.join(ctx[-2:])}"
            else:
                suffixes[e['row']] = base
        else:
            suffixes[e['row']] = base

    # Third pass: resolve collisions iteratively
    e_map = {e['row']: e for e in entries}

    def primary_obj_type(obj_path):
        """Extract primary object type: Radio, SSID, AccessPoint, EndPoint, etc."""
        parts = [p for p in obj_path.rstrip('.').split('.') if p and '{' not in p and p.lower() != 'wifi']
        return sanitize(parts[0]) if parts else 'unknown'

    for _round in range(3):
        suffix_groups = defaultdict(list)
        for row, sfx in suffixes.items():
            suffix_groups[sfx].append(row)

        collisions = {sfx: rows for sfx, rows in suffix_groups.items() if len(rows) > 1}
        if not collisions:
            break

        for sfx, row_list in collisions.items():
            types = {row: primary_obj_type(e_map[row]['obj']) for row in row_list}
            if len(set(types.values())) > 1:
                for row in row_list:
                    suffixes[row] = f"{sfx}_{types[row]}"
                continue

            vals = {}
            for row in row_list:
                _, _, _, hlapi_val = build_suffix(e_map[row])
                vals[row] = hlapi_val
            if len(set(v for v in vals.values() if v)) > 1:
                for row in row_list:
                    if vals[row]:
                        suffixes[row] = f"{sfx}_{vals[row]}"
            else:
                for row in row_list:
                    has_steps = bool(e_map[row]['steps_raw'].strip())
                    suffixes[row] = f"{sfx}_verified" if has_steps else f"{sfx}_basic"

    # Final check
    suffix_groups_final = defaultdict(list)
    for row, sfx in suffixes.items():
        suffix_groups_final[sfx].append(row)
    for sfx, row_list in suffix_groups_final.items():
        if len(row_list) > 1:
            print(f"  WARN: still duplicate suffix '{sfx}' for rows {row_list} (row prefix ensures uniqueness)")

    # ── Build final filenames with row prefix ──
    filenames = {row: f"D{row:03d}_{sfx}" for row, sfx in suffixes.items()}

    # Verify all unique
    all_stems = list(filenames.values())
    assert len(set(all_stems)) == len(all_stems), "Filename collision!"

    # ── Remove old files ──
    removed = sum(1 for f in out_dir.glob("D[0-9]*.yaml") if (f.unlink() or True))
    print(f"Removed {removed} old files")

    # ── Generate YAML ──
    e_map = {e['row']: e for e in entries}
    generated = 0
    for row, stem in sorted(filenames.items()):
        e = e_map[row]
        steps_parsed = parse_steps(e['steps_raw'])
        all_text = e['steps_raw'] + ' ' + e['hlapi']
        needs_sta, needs_ep = detect_topology(all_text)

        devices = {'DUT': {'role': 'ap', 'transport': 'serial'}}
        links = []
        if needs_sta:
            devices['STA'] = {'role':'sta','transport':'adb',
                'config':[{'iface':'5g','mode':'sta','ssid':'{{SSID_5G}}','key':'{{KEY_5G}}'}]}
            links.append({'from':'STA','to':'DUT','band':'5g'})
        if needs_ep:
            devices['EndpointPC'] = {'role':'endpoint','transport':'ssh'}
            links.append({'from':'DUT','to':'EndpointPC','interface':'eth1'})
        topo = {'devices': devices}
        if links: topo['links'] = links

        yaml_steps = []
        if steps_parsed:
            for i, s in enumerate(steps_parsed):
                step = {'id':f'step{i+1}','action':'exec','target':'DUT','command':s}
                if i > 0: step['depends_on'] = f'step{i}'
                if i == len(steps_parsed)-1: step['capture'] = 'result'
                yaml_steps.append(step)
        else:
            cmd = e['hlapi'] if e['hlapi'] else f"echo 'TODO: {e['param']}'"
            yaml_steps = [{'id':'query','action':'exec','target':'DUT','command':cmd,'capture':'result'}]

        criteria = []
        hlapi = e['hlapi']
        m = re.search(r'=\s*(.+)$', hlapi)
        if m:
            exp_val = m.group(1).strip().rstrip(',')
            criteria.append({'field':f"result.{e['param']}",'operator':'contains','value':exp_val,
                'description':f"{e['param']} 應回傳包含 {e['param']}={exp_val}"})
        elif '()' in hlapi:
            criteria.append({'field':'result','operator':'not_contains','value':'error',
                'description':f"{e['param']} 執行不應回傳 error"})
        else:
            criteria.append({'field':'result','operator':'not_empty','value':'',
                'description':f"{e['param']} 應回傳非空結果"})

        case_id = f"wifi-llapi-{stem.replace('_', '-')}"
        case = {
            'id': case_id,
            'name': f"{e['param']} — {e['obj']}",
            'version': '1.0',
            'source': {'report': xlsx_path.name,
                'sheet':'Wifi_LLAPI','row':row,'object':e['obj'],'api':e['param']},
            'platform': {'prplos':'4.0.3','bdk':'6.3.1'},
            'hlapi_command': hlapi,
            'llapi_support': e['llapi_sup'],
            'implemented_by': e['impl_by'],
            'topology': topo,
            'steps': yaml_steps,
            'pass_criteria': criteria,
        }
        if e['verify_cmd']: case['verification_command'] = e['verify_cmd']
        case['results_reference'] = {'v4.0.1':e['r401'],'v4.0.3':e['r403']}

        with open(out_dir / f"{stem}.yaml", 'w') as f:
            yaml.dump(case, f, default_flow_style=False, allow_unicode=True, sort_keys=False, width=120)
        generated += 1

    wb.close()
    print(f"Generated {generated} YAML files, all unique: {len(set(filenames.values()))}/{len(filenames)}")

    # ── Verify: show previously problematic groups ──
    for label, pat in [("getssidstats", "D*getssidstats*"), ("getradiostats", "D*getradiostats*"),
                        ("ac_vo", "D*ac_vo*"), ("enable", "D*_enable_*"),
                        ("discoverymethodenabled", "D*discoverymethodenabled*"),
                        ("channel", "D*_channel_*"), ("retrycount", "D*retrycount*")]:
        files = sorted(out_dir.glob(f"{pat}.yaml"))
        if files:
            print(f"\n--- {label} ({len(files)}) ---")
            for f in files:
                print(f"  {f.name}")


if __name__ == "__main__":
    main()
