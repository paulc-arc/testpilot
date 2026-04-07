from __future__ import annotations

import json
from pathlib import Path

from openpyxl import Workbook
import yaml

from testpilot.reporting.wifi_llapi_compare_0401 import (
    compare_run_against_0401,
    load_0401_answers,
    normalize_result,
    render_compare_markdown,
)
from testpilot.core.case_utils import case_band_results


def _create_answers_xlsx(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Wifi_LLAPI"
    ws["A1"] = "Object"
    ws["C1"] = "Parameter Name"
    ws["G1"] = "Test steps"
    ws["H1"] = "Command Output"
    ws["R1"] = "WiFi 5g"
    ws["S1"] = "WiFi 6g"
    ws["T1"] = "WiFi 2.4g"

    ws["A4"] = "WiFi.AccessPoint.{i}."
    ws["C4"] = "kickStation()"
    ws["G4"] = "kick station steps"
    ws["H4"] = "driver proof"
    ws["R4"] = "Pass"
    ws["S4"] = "Pass"
    ws["T4"] = "Pass"

    for row in range(5, 75):
        ws[f"A{row}"] = f"placeholder-{row}"
        ws[f"C{row}"] = f"placeholder-api-{row}"
        ws[f"R{row}"] = "Fail"
        ws[f"S{row}"] = "Fail"
        ws[f"T{row}"] = "Fail"

    ws["A75"] = "WiFi.AccessPoint.{i}.IEEE80211u."
    ws["C75"] = "InterworkingEnable"
    ws["G75"] = "toggle interworking"
    ws["H75"] = "beacon verification"
    ws["R75"] = "Not Supported"
    ws["S75"] = "Fail"
    ws["T75"] = "skip"

    wb.save(path)
    wb.close()


def _create_case_yaml(path: Path, *, case_id: str, source_row: int, source_object: str, source_api: str) -> None:
    payload = {
        "id": case_id,
        "name": case_id,
        "source": {
            "row": source_row,
            "object": source_object,
            "api": source_api,
            "baseline": "v4.0.3",
        },
        "results_reference": {
            "v4.0.3": {
                "5g": "Pass",
                "6g": "Pass",
                "2.4g": "Pass",
            }
        },
    }
    path.write_text(yaml.safe_dump(payload, sort_keys=False), encoding="utf-8")


def _create_trace(path: Path, *, case_id: str, source_row: int, evaluation_verdict: str, status: str) -> None:
    payload = {
        "case_id": case_id,
        "source_row": source_row,
        "final": {
            "status": status,
            "evaluation_verdict": evaluation_verdict,
            "attempts_used": 1,
            "comment": "",
        },
        "attempts": [],
    }
    path.write_text(json.dumps(payload), encoding="utf-8")


def test_normalize_result_only_pass_stays_pass():
    assert normalize_result("Pass") == "Pass"
    assert normalize_result(" pass ") == "Pass"
    assert normalize_result("Fail") == "Fail"
    assert normalize_result("Not Supported") == "Fail"
    assert normalize_result("skip") == "Fail"
    assert normalize_result("") == "Fail"


def test_load_0401_answers_reads_rst_columns(tmp_path: Path):
    answers_xlsx = tmp_path / "0401.xlsx"
    _create_answers_xlsx(answers_xlsx)

    answers = load_0401_answers(answers_xlsx)

    assert answers[4].api == "kickStation()"
    assert answers[4].norm_5g == "Pass"
    assert answers[75].norm_5g == "Fail"
    assert answers[75].norm_6g == "Fail"
    assert answers[75].norm_24g == "Fail"


def test_compare_run_against_0401_uses_dnum_row_mapping(tmp_path: Path):
    answers_xlsx = tmp_path / "0401.xlsx"
    _create_answers_xlsx(answers_xlsx)

    cases_dir = tmp_path / "cases"
    cases_dir.mkdir()
    _create_case_yaml(
        cases_dir / "D004_kickstation.yaml",
        case_id="wifi-llapi-D004-kickstation",
        source_row=2,
        source_object="WiFi.AccessPoint.{i}.",
        source_api="kickStation()",
    )
    _create_case_yaml(
        cases_dir / "D075_interworkingenable.yaml",
        case_id="wifi-llapi-D075-interworkingenable-accesspoint",
        source_row=77,
        source_object="WiFi.AccessPoint.{i}.IEEE80211u",
        source_api="InterworkingEnable",
    )

    trace_dir = tmp_path / "agent_trace"
    trace_dir.mkdir()
    _create_trace(
        trace_dir / "wifi-llapi-D004-kickstation.json",
        case_id="wifi-llapi-D004-kickstation",
        source_row=2,
        evaluation_verdict="Pass",
        status="Pass",
    )
    _create_trace(
        trace_dir / "wifi-llapi-D075-interworkingenable-accesspoint.json",
        case_id="wifi-llapi-D075-interworkingenable-accesspoint",
        source_row=77,
        evaluation_verdict="Fail",
        status="Fail",
    )

    payload = compare_run_against_0401(trace_dir, answers_xlsx, cases_dir=cases_dir)

    assert payload["compared_case_count"] == 2
    assert payload["full_match_count"] == 2
    assert payload["mismatch_case_count"] == 0
    assert payload["metadata_drift_count"] == 0


def test_render_compare_markdown_lists_mismatches(tmp_path: Path):
    answers_xlsx = tmp_path / "0401.xlsx"
    _create_answers_xlsx(answers_xlsx)

    cases_dir = tmp_path / "cases"
    cases_dir.mkdir()
    _create_case_yaml(
        cases_dir / "D004_kickstation.yaml",
        case_id="wifi-llapi-D004-kickstation",
        source_row=2,
        source_object="WiFi.AccessPoint.{i}.",
        source_api="kickStation()",
    )

    trace_dir = tmp_path / "agent_trace"
    trace_dir.mkdir()
    _create_trace(
        trace_dir / "wifi-llapi-D004-kickstation.json",
        case_id="wifi-llapi-D004-kickstation",
        source_row=2,
        evaluation_verdict="Fail",
        status="Fail",
    )

    payload = compare_run_against_0401(trace_dir, answers_xlsx, cases_dir=cases_dir)
    md = render_compare_markdown(payload)

    assert payload["mismatch_case_count"] == 1
    assert "wifi-llapi-D004-kickstation" in md
    assert "mismatch bands" in md.lower()
    assert "expected raw" in md.lower()


def test_compare_run_against_0401_overlays_later_runs(tmp_path: Path):
    answers_xlsx = tmp_path / "0401.xlsx"
    _create_answers_xlsx(answers_xlsx)

    cases_dir = tmp_path / "cases"
    cases_dir.mkdir()
    _create_case_yaml(
        cases_dir / "D004_kickstation.yaml",
        case_id="wifi-llapi-D004-kickstation",
        source_row=2,
        source_object="WiFi.AccessPoint.{i}.",
        source_api="kickStation()",
    )

    full_run_dir = tmp_path / "full_run"
    full_run_dir.mkdir()
    _create_trace(
        full_run_dir / "wifi-llapi-D004-kickstation.json",
        case_id="wifi-llapi-D004-kickstation",
        source_row=2,
        evaluation_verdict="Fail",
        status="Fail",
    )

    rerun_dir = tmp_path / "rerun"
    rerun_dir.mkdir()
    _create_trace(
        rerun_dir / "wifi-llapi-D004-kickstation.json",
        case_id="wifi-llapi-D004-kickstation",
        source_row=2,
        evaluation_verdict="Pass",
        status="Pass",
    )

    payload = compare_run_against_0401(
        [full_run_dir, rerun_dir],
        answers_xlsx,
        cases_dir=cases_dir,
    )

    assert payload["trace_dirs"] == [str(full_run_dir), str(rerun_dir)]
    assert payload["full_match_count"] == 1
    assert payload["mismatch_case_count"] == 0
    assert payload["cases"][0]["trace_path"] == str(rerun_dir / "wifi-llapi-D004-kickstation.json")


def test_case_band_results_falls_back_to_v403_without_source_baseline():
    case = {
        "results_reference": {
            "v4.0.1": {"5g": "Fail", "6g": "Fail", "2.4g": "Fail"},
            "v4.0.3": {"5g": "Pass", "6g": "Blocker", "2.4g": "Pass"},
        },
        "source": {
            "row": 4,
            "object": "WiFi.AccessPoint.{i}.",
            "api": "sendBssTransferRequest()",
        },
    }

    assert case_band_results(case, True) == ("Pass", "Blocker", "Pass")
