"""Integration tests for per-case agent dispatch behavior in wifi_llapi run path."""

from __future__ import annotations

import json
from pathlib import Path
from typing import Any

from openpyxl import Workbook

from testpilot.core.orchestrator import Orchestrator

FAIL_CASE_ID = "wifi-llapi-D004-retry-fail"
PASS_CASE_ID = "wifi-llapi-D005-pass-after-fail"
FAIL_CASE_LEGACY_ID = "wifi-llapi-r004-retry-fail"
PASS_CASE_LEGACY_ID = "wifi-llapi-r005-pass-after-fail"


def _write_testbed_yaml(path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(
        """
testbed:
  name: unit-testbed
  devices:
    DUT:
      role: ap
""".strip()
        + "\n",
        encoding="utf-8",
    )


def _write_source_xlsx(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Wifi_LLAPI"

    ws["A2"] = "Object"
    ws["C2"] = "Parameter Name"

    ws["A4"] = "WiFi.AccessPoint.{i}."
    ws["C4"] = "kickStation()"
    ws["A5"] = "WiFi.Radio.{i}."
    ws["C5"] = "scan()"

    wb.save(path)
    wb.close()


def _write_wifi_llapi_plugin(plugin_dir: Path) -> None:
    plugin_dir.mkdir(parents=True, exist_ok=True)
    (plugin_dir / "plugin.py").write_text(
        """
from __future__ import annotations

from pathlib import Path
from typing import Any

from testpilot.core.plugin_base import PluginBase


CASES = [
    {
        "id": "wifi-llapi-D004-retry-fail",
        "aliases": ["wifi-llapi-r004-retry-fail"],
        "name": "retry-fail",
        "source": {
            "row": 4,
            "object": "WiFi.AccessPoint.{i}.",
            "api": "kickStation()",
        },
        "steps": [{"id": "s1", "command": "ubus-cli fail"}],
        "simulation": "always_fail",
        "bands": ["5g"],
    },
    {
        "id": "wifi-llapi-D005-pass-after-fail",
        "aliases": ["wifi-llapi-r005-pass-after-fail"],
        "name": "pass-after-fail",
        "source": {
            "row": 5,
            "object": "WiFi.Radio.{i}.",
            "api": "scan()",
        },
        "steps": [{"id": "s1", "command": "ubus-cli pass"}],
        "simulation": "always_pass",
        "bands": ["6g"],
    },
]


class Plugin(PluginBase):
    def __init__(self) -> None:
        self.execute_calls: dict[str, int] = {}

    @property
    def name(self) -> str:
        return "wifi_llapi"

    @property
    def version(self) -> str:
        return "0.0-test"

    @property
    def cases_dir(self) -> Path:
        return Path(__file__).parent / "cases"

    def discover_cases(self) -> list[dict[str, Any]]:
        return [dict(c) for c in CASES]

    def setup_env(self, case: dict[str, Any], topology: Any) -> bool:
        return True

    def verify_env(self, case: dict[str, Any], topology: Any) -> bool:
        return True

    def execute_step(self, case: dict[str, Any], step: dict[str, Any], topology: Any) -> dict[str, Any]:
        case_id = str(case.get("id", ""))
        self.execute_calls[case_id] = self.execute_calls.get(case_id, 0) + 1
        attempt = self.execute_calls[case_id]
        should_fail = case.get("simulation") == "always_fail"
        return {
            "success": not should_fail,
            "output": f"{case_id}:attempt={attempt}",
            "captured": {"attempt": attempt},
            "timing": 0.01,
        }

    def evaluate(self, case: dict[str, Any], results: dict[str, Any]) -> bool:
        steps = results.get("steps", {})
        if not isinstance(steps, dict):
            return False
        return all(bool(item.get("success", False)) for item in steps.values())

    def teardown(self, case: dict[str, Any], topology: Any) -> None:
        return None
""".strip()
        + "\n",
        encoding="utf-8",
    )

    (plugin_dir / "agent-config.yaml").write_text(
        """
version: 1
default_mode: headless
selection_policy:
  fallback: automatic
  on_unavailable: next_priority
execution:
  scope: per_case
  mode: sequential
  max_concurrency: 1
  failure_policy: retry_then_fail_and_continue
  retry:
    max_attempts: 2
    backoff_seconds: 0
  timeout:
    base_seconds: 12
    per_step_seconds: 3
    retry_multiplier: 2.0
    max_seconds: 60
runners:
  - priority: 1
    cli_agent: unavailable-runner
    model: missing-model
    effort: high
    enabled: true
  - priority: 2
    cli_agent: fallback-runner
    model: fallback-model
    effort: high
    enabled: true
""".strip()
        + "\n",
        encoding="utf-8",
    )


def _build_temp_project(tmp_path: Path) -> tuple[Path, Path]:
    root = tmp_path / "project"
    _write_testbed_yaml(root / "configs" / "testbed.yaml")
    _write_wifi_llapi_plugin(root / "plugins" / "wifi_llapi")

    source_xlsx = root / "source.xlsx"
    _write_source_xlsx(source_xlsx)
    return root, source_xlsx


def _run_wifi_llapi(tmp_path: Path) -> tuple[dict[str, Any], Orchestrator, Path]:
    project_root, source_xlsx = _build_temp_project(tmp_path)

    orch = Orchestrator(
        project_root=project_root,
        plugins_dir=project_root / "plugins",
        config_path=project_root / "configs" / "testbed.yaml",
    )
    result = orch.run(
        "wifi_llapi",
        case_ids=[FAIL_CASE_LEGACY_ID, PASS_CASE_LEGACY_ID],
        dut_fw_ver="FW-TEST-1",
        report_source_xlsx=str(source_xlsx),
    )
    return result, orch, project_root


def _iter_dict_nodes(node: Any):
    if isinstance(node, dict):
        yield node
        for value in node.values():
            yield from _iter_dict_nodes(value)
        return
    if isinstance(node, list):
        for item in node:
            yield from _iter_dict_nodes(item)


def _to_int(value: Any) -> int | None:
    try:
        return int(value)
    except (TypeError, ValueError):
        return None


def _to_float(value: Any) -> float | None:
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _normalize_runner(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, str):
        return value
    if isinstance(value, dict):
        for key in ("cli_agent", "runner", "name", "id", "model"):
            candidate = value.get(key)
            if isinstance(candidate, str) and candidate:
                return candidate
        return json.dumps(value, ensure_ascii=False, sort_keys=True)
    return str(value)


def _extract_trace_events(payload: Any) -> list[dict[str, Any]]:
    events: list[dict[str, Any]] = []
    root_case_id = payload.get("case_id") if isinstance(payload, dict) else None
    for node in _iter_dict_nodes(payload):
        case_id = node.get("case_id", root_case_id)
        if case_id is None and isinstance(node.get("case"), dict):
            case_id = node["case"].get("id")

        attempt = _to_int(
            node.get("attempt", node.get("attempt_index", node.get("retry_attempt")))
        )

        runner_raw = node.get(
            "runner",
            node.get(
                "selected_runner",
                node.get("cli_agent", node.get("agent")),
            ),
        )
        runner = _normalize_runner(runner_raw)

        timeout = _to_float(
            node.get(
                "timeout",
                node.get("attempt_timeout", node.get("timeout_seconds")),
            )
        )

        if isinstance(case_id, str) and attempt is not None and runner and timeout is not None:
            events.append(
                {
                    "case_id": case_id,
                    "attempt": attempt,
                    "runner": runner,
                    "timeout": timeout,
                }
            )
    return events


def _load_case_traces(
    trace_root: Path,
    expected_case_ids: set[str],
) -> dict[str, list[tuple[Path, Any]]]:
    traces: dict[str, list[tuple[Path, Any]]] = {case_id: [] for case_id in expected_case_ids}

    for path in sorted(trace_root.rglob("*.json")):
        try:
            payload = json.loads(path.read_text(encoding="utf-8"))
        except json.JSONDecodeError:
            continue

        matched: set[str] = set()

        if isinstance(payload, dict):
            case_id = payload.get("case_id")
            if isinstance(case_id, str) and case_id in expected_case_ids:
                matched.add(case_id)

        lower_name = path.name.lower()
        for case_id in expected_case_ids:
            if case_id.lower() in lower_name:
                matched.add(case_id)

        for event in _extract_trace_events(payload):
            case_id = event["case_id"]
            if case_id in expected_case_ids:
                matched.add(case_id)

        for case_id in matched:
            traces[case_id].append((path, payload))

    return traces


def test_retry_then_fail_and_continue_and_summary(tmp_path: Path):
    result, orch, _project_root = _run_wifi_llapi(tmp_path)

    assert result["status"] == "completed"
    assert result["pass_count"] == 1
    assert result["fail_count"] == 1
    assert Path(result["report_path"]).is_file()

    plugin = orch.loader.load("wifi_llapi")
    execute_calls = getattr(plugin, "execute_calls", {})

    # retry_then_fail_and_continue: first case should be retried,
    # then next case should still execute.
    assert execute_calls.get(FAIL_CASE_ID, 0) >= 2
    assert execute_calls.get(PASS_CASE_ID, 0) >= 1


def test_fallback_trace_and_per_case_trace_payload(tmp_path: Path, monkeypatch):
    monkeypatch.setenv(
        "TESTPILOT_AGENT_AVAILABILITY",
        json.dumps(
            {
                "wifi_llapi": {
                    "unavailable-runner": "binary_not_found",
                    "fallback-runner": True,
                }
            }
        ),
    )
    _result, _orch, project_root = _run_wifi_llapi(tmp_path)

    trace_root = project_root / "plugins" / "wifi_llapi" / "reports" / "agent_trace"
    assert trace_root.is_dir(), f"missing per-case trace root: {trace_root}"

    expected_case_ids = {FAIL_CASE_ID, PASS_CASE_ID}
    traces = _load_case_traces(trace_root, expected_case_ids)

    for case_id in expected_case_ids:
        assert traces[case_id], f"missing per-case trace file for {case_id}"

        case_events: list[dict[str, Any]] = []
        for _path, payload in traces[case_id]:
            case_events.extend(
                [event for event in _extract_trace_events(payload) if event["case_id"] == case_id]
            )

        assert case_events, f"trace payload missing case_id/attempt/runner/timeout for {case_id}"
        assert all(event["attempt"] >= 1 for event in case_events)
        assert all(event["runner"] for event in case_events)
        assert all(event["timeout"] > 0 for event in case_events)

        selection_trace = traces[case_id][0][1].get("selection_trace", {})
        selected = selection_trace.get("selected", {})
        selected_runner = _normalize_runner(selected)
        fallback = selection_trace.get("fallback", {})

        assert selected_runner == "fallback-runner"
        assert isinstance(fallback, dict) and fallback.get("applied") is True
        reason = str(fallback.get("reason", "")).lower()
        assert any(token in reason for token in ("unavailable", "fallback", "next_priority"))

    fail_events: list[dict[str, Any]] = []
    for _path, payload in traces[FAIL_CASE_ID]:
        fail_events.extend(
            [event for event in _extract_trace_events(payload) if event["case_id"] == FAIL_CASE_ID]
        )

    attempts = sorted({event["attempt"] for event in fail_events})
    assert attempts and attempts[0] == 1
    assert attempts[-1] >= 2

    timeout_by_attempt: dict[int, float] = {}
    for event in fail_events:
        timeout_by_attempt.setdefault(event["attempt"], event["timeout"])

    assert 1 in timeout_by_attempt
    assert 2 in timeout_by_attempt
    assert timeout_by_attempt[2] > timeout_by_attempt[1]

    trace_text = " ".join(
        json.dumps(payload, ensure_ascii=False).lower()
        for case_entries in traces.values()
        for _path, payload in case_entries
    )
    assert "unavailable-runner" in trace_text
