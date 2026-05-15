"""Tests for CLI entry points — verifies all commands are wired and produce expected output."""

from __future__ import annotations

import re
import sys
import textwrap
from importlib.util import module_from_spec, spec_from_file_location
import json
from pathlib import Path
from typing import Any

from openpyxl import Workbook

from click.testing import CliRunner

import testpilot.cli
from testpilot.core.testbed_config import TestbedConfig
from testpilot.cli import main


_DUMMY_PLUGIN_PY = textwrap.dedent(
    """
    from __future__ import annotations
    from pathlib import Path
    from typing import Any
    from testpilot.core.plugin_base import PluginBase

    class Plugin(PluginBase):
        @property
        def name(self) -> str:
            return "stage_marker"

        @property
        def cases_dir(self) -> Path:
            return Path(__file__).parent / "cases"

        def discover_cases(self) -> list[dict[str, Any]]:
            return []

        def execute_step(self, case, step, topology):
            return {"success": True, "output": "", "captured": {}, "timing": 0.0}

        def evaluate(self, case, results):
            return True
    """
).lstrip()


def _make_stage_marker_plugin(root: Path, plugin_name: str, marker: str) -> None:
    plugin_dir = root / "plugins" / plugin_name
    plugin_dir.mkdir(parents=True)
    (plugin_dir / "plugin.py").write_text(_DUMMY_PLUGIN_PY, encoding="utf-8")
    (plugin_dir / "testbed.yaml.example").write_text(
        f"testbed:\n  name: {marker}\n", encoding="utf-8"
    )
    (root / "configs").mkdir(exist_ok=True)


def _clear_provider_env(monkeypatch) -> None:
    for key in (
        "COPILOT_PROVIDER_TYPE",
        "COPILOT_PROVIDER_BASE_URL",
        "COPILOT_PROVIDER_API_KEY",
        "COPILOT_MODEL",
        "COPILOT_PROVIDER_AZURE_API_VERSION",
    ):
        monkeypatch.delenv(key, raising=False)


def _load_real_brcm_plugin() -> Any:
    root = Path(__file__).resolve().parents[1]
    plugin_path = root / "plugins" / "brcm_fw_upgrade" / "plugin.py"
    spec = spec_from_file_location("tests_brcm_fw_upgrade_plugin_cli", plugin_path)
    assert spec is not None
    assert spec.loader is not None
    added_to_path = str(root) not in sys.path
    if added_to_path:
        sys.path.insert(0, str(root))
    try:
        module = module_from_spec(spec)
        spec.loader.exec_module(module)
    finally:
        if added_to_path:
            sys.path.remove(str(root))
    return module.Plugin()


def test_version_flag():
    """--version prints version string."""
    runner = CliRunner()
    result = runner.invoke(main, ["--version"])
    assert result.exit_code == 0
    assert "testpilot" in result.output.lower()


def test_list_plugins_shows_wifi_llapi():
    """list-plugins discovers wifi_llapi."""
    runner = CliRunner()
    result = runner.invoke(main, ["list-plugins"])
    assert result.exit_code == 0
    assert "wifi_llapi" in result.output


def test_list_cases_for_wifi_llapi():
    """list-cases wifi_llapi returns non-empty list."""
    runner = CliRunner()
    result = runner.invoke(main, ["list-cases", "wifi_llapi"])
    assert result.exit_code == 0
    # Should contain at least one D### case ID
    assert "D0" in result.output or "D1" in result.output


def test_list_cases_unknown_plugin():
    """list-cases with unknown plugin name raises error."""
    runner = CliRunner()
    result = runner.invoke(main, ["list-cases", "nonexistent_plugin"])
    assert result.exit_code != 0


def test_run_without_dut_fw_ver_uses_default(monkeypatch):
    """run command accepts default --dut-fw-ver without crashing on missing transport."""
    _clear_provider_env(monkeypatch)
    calls: list[dict[str, Any]] = []

    class FakeOrchestrator:
        def __init__(self, project_root):
            self.project_root = project_root

        def run(self, plugin_name, case_ids, **kwargs):
            calls.append(
                {
                    "plugin_name": plugin_name,
                    "case_ids": case_ids,
                    "kwargs": kwargs,
                }
            )
            return {"status": "ok", "plugin": plugin_name, "case_ids": case_ids}

    monkeypatch.setattr(testpilot.cli, "Orchestrator", FakeOrchestrator)

    runner = CliRunner()
    result = runner.invoke(main, ["run", "wifi_llapi", "--case", "wifi-llapi-D004-kickstation"])

    assert result.exit_code == 0
    assert calls == [
        {
            "plugin_name": "wifi_llapi",
            "case_ids": ["wifi-llapi-D004-kickstation"],
            "kwargs": {
                "dut_fw_ver": "DUT-FW-VER",
                "provider_config": None,
            },
        }
    ]


def test_wifi_llapi_command_uses_same_run_path(monkeypatch):
    """wifi_llapi command invokes the same normal run path with wifi_llapi plugin."""
    _clear_provider_env(monkeypatch)
    calls: list[dict[str, Any]] = []

    class FakeOrchestrator:
        def __init__(self, project_root):
            self.project_root = project_root

        def run(self, plugin_name, case_ids, **kwargs):
            calls.append(
                {
                    "plugin_name": plugin_name,
                    "case_ids": case_ids,
                    "kwargs": kwargs,
                }
            )
            return {"status": "ok", "plugin": plugin_name, "case_ids": case_ids}

    monkeypatch.setattr(testpilot.cli, "Orchestrator", FakeOrchestrator)

    runner = CliRunner()
    result = runner.invoke(
        main,
        [
            "wifi_llapi",
            "--case",
            "wifi-llapi-D004-kickstation",
            "--dut-fw-ver",
            "BGW720-B0-403",
        ],
    )

    assert result.exit_code == 0
    assert calls == [
        {
            "plugin_name": "wifi_llapi",
            "case_ids": ["wifi-llapi-D004-kickstation"],
            "kwargs": {
                "dut_fw_ver": "BGW720-B0-403",
                "provider_config": None,
            },
        }
    ]


def test_wifi_llapi_help_shows_operational_usage():
    """wifi_llapi help shows the fixed QC/TEST operational command format."""
    runner = CliRunner()

    result = runner.invoke(main, ["wifi_llapi", "--help"])

    assert result.exit_code == 0
    assert "testpilot wifi_llapi [--case CASE_ID] [--dut-fw-ver FW_VER]" in result.output


def test_wifi_llapi_without_case_runs_all_cases(monkeypatch):
    """wifi_llapi without --case passes case_ids=None to run all cases."""
    _clear_provider_env(monkeypatch)
    calls: list[dict[str, Any]] = []

    class FakeOrchestrator:
        def __init__(self, project_root):
            self.project_root = project_root

        def run(self, plugin_name, case_ids, **kwargs):
            calls.append(
                {
                    "plugin_name": plugin_name,
                    "case_ids": case_ids,
                    "kwargs": kwargs,
                }
            )
            return {"status": "ok", "plugin": plugin_name, "case_ids": case_ids}

    monkeypatch.setattr(testpilot.cli, "Orchestrator", FakeOrchestrator)

    runner = CliRunner()
    result = runner.invoke(main, ["wifi_llapi", "--dut-fw-ver", "BGW720-B0-403"])

    assert result.exit_code == 0
    assert calls == [
        {
            "plugin_name": "wifi_llapi",
            "case_ids": None,
            "kwargs": {
                "dut_fw_ver": "BGW720-B0-403",
                "provider_config": None,
            },
        }
    ]


def test_run_wifi_llapi_rejects_removed_report_source_xlsx_flag(monkeypatch):
    _clear_provider_env(monkeypatch)
    runner = CliRunner()

    result = runner.invoke(
        main,
        [
            "run",
            "wifi_llapi",
            "--report-source-xlsx",
            "0401.xlsx",
        ],
    )

    assert result.exit_code != 0
    assert "No such option: --report-source-xlsx" in result.output


def test_run_without_plugin_name_shows_correct_format_guidance():
    """run without plugin_name returns a targeted message with correct syntax."""
    runner = CliRunner()
    result = runner.invoke(main, ["run", "--case", "wifi-llapi-D004-kickstation"])

    assert result.exit_code != 0
    assert "Missing required argument PLUGIN_NAME." in result.output
    assert "testpilot run <plugin_name> [--case <case_id>]" in result.output
    assert "testpilot run wifi_llapi --case wifi-llapi-D004-kickstation" in result.output
    assert "testpilot list-cases wifi_llapi" in result.output


def test_help_text_for_main():
    """Main --help shows all subcommands."""
    runner = CliRunner()
    result = runner.invoke(main, ["--help"])
    assert result.exit_code == 0
    assert "list-plugins" in result.output
    assert "list-cases" in result.output
    assert "run" in result.output
    assert "brcm-fw-upgrade" in result.output
    assert "wifi-llapi" in result.output


def test_help_text_for_wifi_llapi_group():
    """wifi-llapi --help shows subcommands."""
    runner = CliRunner()
    result = runner.invoke(main, ["wifi-llapi", "--help"])
    assert result.exit_code == 0
    assert "baseline-qualify" in result.output
    assert "build-template-report" in result.output
    assert "audit-yaml-commands" in result.output


def test_help_text_for_brcm_fw_upgrade_group():
    runner = CliRunner()
    result = runner.invoke(main, ["brcm-fw-upgrade", "--help"])
    assert result.exit_code == 0
    assert "run" in result.output


def test_brcm_fw_upgrade_run_invokes_plugin_with_runtime_overrides(monkeypatch):
    _clear_provider_env(monkeypatch)
    calls: list[dict[str, Any]] = []

    class FakePlugin:
        def run_cases(self, topology, *, case_ids, runtime_overrides):
            calls.append(
                {
                    "topology": topology,
                    "case_ids": case_ids,
                    "runtime_overrides": runtime_overrides,
                }
            )
            return {"status": "ok", "case_ids": case_ids}

    class FakeLoader:
        def load(self, plugin_name: str):
            assert plugin_name == "brcm_fw_upgrade"
            return FakePlugin()

    class FakeOrchestrator:
        def __init__(self, project_root):
            self.project_root = project_root
            self.loader = FakeLoader()
            self.config = {"name": "fake-topology"}

    monkeypatch.setattr(testpilot.cli, "Orchestrator", FakeOrchestrator)

    runner = CliRunner()
    result = runner.invoke(
        main,
        [
            "brcm-fw-upgrade",
            "run",
            "--case",
            "brcm-fw-upgrade-dut-sta-forward",
            "--forward-image",
            "/tmp/0410.pkgtb",
            "--rollback-image",
            "/tmp/0403.pkgtb",
            "--fw-name",
            "bcmBGW720-300_squashfs_full_update.pkgtb",
            "--expected-image-tag",
            "631BGW720-3001101323",
            "--expected-build-time",
            "Mon Apr 20 13:02:57 CST 2026",
            "--platform-profile",
            "bgw720_prpl",
            "--topology",
            "dut_plus_sta",
            "--set",
            "DUT_IP=192.168.1.1",
        ],
    )

    assert result.exit_code == 0
    assert calls == [
        {
            "topology": {"name": "fake-topology"},
            "case_ids": ["brcm-fw-upgrade-dut-sta-forward"],
            "runtime_overrides": {
                "FW_FORWARD_PATH": "/tmp/0410.pkgtb",
                "FW_ROLLBACK_PATH": "/tmp/0403.pkgtb",
                "FW_NAME": "bcmBGW720-300_squashfs_full_update.pkgtb",
                "EXPECTED_IMAGE_TAG": "631BGW720-3001101323",
                "EXPECTED_BUILD_TIME": "Mon Apr 20 13:02:57 CST 2026",
                "PLATFORM_PROFILE": "bgw720_prpl",
                "TOPOLOGY": "dut_plus_sta",
                "DUT_IP": "192.168.1.1",
            },
        }
    ]


def test_brcm_fw_upgrade_run_derives_defaults_from_forward_image(monkeypatch, tmp_path: Path):
    _clear_provider_env(monkeypatch)
    calls: list[dict[str, Any]] = []
    forward_image = tmp_path / "bcmBGW720-300_squashfs_full_update.pkgtb"
    forward_image.write_bytes(
        b"header\x00"
        b"@(#) $imageversion: 631BGW720-3001101323 $\x00"
        b"#1 SMP PREEMPT Mon Apr 20 13:02:57 CST 2026\x00"
    )

    class FakePlugin:
        def run_cases(self, topology, *, case_ids, runtime_overrides):
            calls.append(
                {
                    "topology": topology,
                    "case_ids": case_ids,
                    "runtime_overrides": runtime_overrides,
                }
            )
            return {"status": "ok", "selected_case_ids": case_ids, "results": []}

    class FakeLoader:
        def load(self, plugin_name: str):
            assert plugin_name == "brcm_fw_upgrade"
            return FakePlugin()

    class FakeOrchestrator:
        def __init__(self, project_root):
            self.project_root = project_root
            self.loader = FakeLoader()
            self.config = {"name": "fake-topology"}

    monkeypatch.setattr(testpilot.cli, "Orchestrator", FakeOrchestrator)

    runner = CliRunner()
    result = runner.invoke(
        main,
        [
            "brcm-fw-upgrade",
            "run",
            "--case",
            "brcm-fw-upgrade-single-dut-forward",
            "--forward-image",
            str(forward_image),
        ],
    )

    assert result.exit_code == 0
    assert calls == [
        {
            "topology": {"name": "fake-topology"},
            "case_ids": ["brcm-fw-upgrade-single-dut-forward"],
            "runtime_overrides": {
                "FW_FORWARD_PATH": str(forward_image),
                "FW_ROLLBACK_PATH": str(forward_image),
                "FW_NAME": "bcmBGW720-300_squashfs_full_update.pkgtb",
                "EXPECTED_IMAGE_TAG": "631BGW720-3001101323",
                "EXPECTED_BUILD_TIME": "Apr 20 13:02:57 CST 2026",
            },
        }
    ]


def test_brcm_fw_upgrade_run_reports_clean_error_when_metadata_cannot_be_derived(
    monkeypatch, tmp_path: Path
):
    _clear_provider_env(monkeypatch)
    forward_image = tmp_path / "bad.pkgtb"
    forward_image.write_bytes(b"no metadata here")

    class FakePlugin:
        def run_cases(self, topology, *, case_ids, runtime_overrides):
            raise AssertionError("run_cases should not be called when metadata derivation fails")

    class FakeLoader:
        def load(self, plugin_name: str):
            assert plugin_name == "brcm_fw_upgrade"
            return FakePlugin()

    class FakeOrchestrator:
        def __init__(self, project_root):
            self.project_root = project_root
            self.loader = FakeLoader()
            self.config = {"name": "fake-topology"}

    monkeypatch.setattr(testpilot.cli, "Orchestrator", FakeOrchestrator)

    runner = CliRunner()
    result = runner.invoke(
        main,
        [
            "brcm-fw-upgrade",
            "run",
            "--case",
            "brcm-fw-upgrade-single-dut-forward",
            "--forward-image",
            str(forward_image),
        ],
    )

    assert result.exit_code != 0
    assert "failed to derive forward image metadata" in result.output
    assert "Error:" in result.output


def test_brcm_fw_upgrade_run_exits_nonzero_when_plugin_reports_failure(monkeypatch):
    _clear_provider_env(monkeypatch)

    class FakePlugin:
        def run_cases(self, topology, *, case_ids, runtime_overrides):
            return {
                "status": "failed",
                "selected_case_ids": case_ids,
                "results": [{"case_id": case_ids[0], "verdict": False, "comment": "flash failed"}],
            }

    class FakeLoader:
        def load(self, plugin_name: str):
            assert plugin_name == "brcm_fw_upgrade"
            return FakePlugin()

    class FakeOrchestrator:
        def __init__(self, project_root):
            self.project_root = project_root
            self.loader = FakeLoader()
            self.config = {"name": "fake-topology"}

    monkeypatch.setattr(testpilot.cli, "Orchestrator", FakeOrchestrator)

    runner = CliRunner()
    result = runner.invoke(
        main,
        [
            "brcm-fw-upgrade",
            "run",
            "--case",
            "brcm-fw-upgrade-dut-sta-forward",
            "--forward-image",
            "/tmp/0410.pkgtb",
            "--rollback-image",
            "/tmp/0403.pkgtb",
            "--fw-name",
            "bcmBGW720-300_squashfs_full_update.pkgtb",
            "--expected-image-tag",
            "631BGW720-3001101323",
            "--expected-build-time",
            "Mon Apr 20 13:02:57 CST 2026",
            "--platform-profile",
            "bgw720_prpl",
            "--topology",
            "dut_plus_sta",
        ],
    )

    assert result.exit_code == 1
    assert '"status":"failed"' in result.output.replace(" ", "")


def test_brcm_fw_upgrade_run_uses_testbed_config_name(monkeypatch):
    _clear_provider_env(monkeypatch)
    root = Path(__file__).resolve().parents[1]
    plugin = _load_real_brcm_plugin()
    forward_image = Path(__file__).resolve()
    rollback_image = root / "plugins" / "brcm_fw_upgrade" / "plugin.py"

    def _fake_open_live_shells(*, topology, case_topology, profile, fw_name):
        return {}, {}

    def _fake_run_case(*, case_id, shells, runtime_overrides):
        return {"verdict": True, "comment": "", "evidence": {"phases": []}}

    monkeypatch.setattr(plugin, "_open_live_shells", _fake_open_live_shells)
    monkeypatch.setattr(plugin, "run_case", _fake_run_case)

    class FakeLoader:
        def load(self, plugin_name: str):
            assert plugin_name == "brcm_fw_upgrade"
            return plugin

    class FakeOrchestrator:
        def __init__(self, project_root):
            self.project_root = project_root
            self.loader = FakeLoader()
            self.config = TestbedConfig(root / "configs" / "testbed.yaml")

    monkeypatch.setattr(testpilot.cli, "Orchestrator", FakeOrchestrator)

    runner = CliRunner()
    result = runner.invoke(
        main,
        [
            "brcm-fw-upgrade",
            "run",
            "--case",
            "brcm-fw-upgrade-dut-sta-forward",
            "--forward-image",
            str(forward_image),
            "--rollback-image",
            str(rollback_image),
            "--fw-name",
            forward_image.name,
            "--expected-image-tag",
            "631BGW720-3001101323",
            "--expected-build-time",
            "Apr 20 13:02:57 CST 2026",
            "--platform-profile",
            "bgw720_prpl",
            "--topology",
            "dut_plus_sta",
        ],
    )

    assert result.exit_code == 0
    payload = json.loads(result.output)
    assert payload["results"][0]["topology_name"] == "lab-bench-1"


def test_brcm_fw_upgrade_run_rejects_platform_profile_mismatch(monkeypatch):
    _clear_provider_env(monkeypatch)
    root = Path(__file__).resolve().parents[1]

    class FakeLoader:
        def load(self, plugin_name: str):
            assert plugin_name == "brcm_fw_upgrade"
            return _load_real_brcm_plugin()

    class FakeOrchestrator:
        def __init__(self, project_root):
            self.project_root = project_root
            self.loader = FakeLoader()
            self.config = TestbedConfig(root / "configs" / "testbed.yaml")

    monkeypatch.setattr(testpilot.cli, "Orchestrator", FakeOrchestrator)

    runner = CliRunner()
    result = runner.invoke(
        main,
        [
            "brcm-fw-upgrade",
            "run",
            "--case",
            "brcm-fw-upgrade-dut-sta-forward",
            "--forward-image",
            "images/0410.pkgtb",
            "--rollback-image",
            "images/0403.pkgtb",
            "--fw-name",
            "bcmBGW720-300_squashfs_full_update.pkgtb",
            "--expected-image-tag",
            "631BGW720-3001101323",
            "--expected-build-time",
            "Mon Apr 20 13:02:57 CST 2026",
            "--platform-profile",
            "wrong_profile",
            "--topology",
            "dut_plus_sta",
        ],
    )

    assert result.exit_code != 0
    assert "requires platform_profile 'bgw720_prpl', got 'wrong_profile'" in result.output
    assert "Error:" in result.output


def test_brcm_fw_upgrade_run_rejects_topology_mismatch(monkeypatch):
    _clear_provider_env(monkeypatch)
    root = Path(__file__).resolve().parents[1]

    class FakeLoader:
        def load(self, plugin_name: str):
            assert plugin_name == "brcm_fw_upgrade"
            return _load_real_brcm_plugin()

    class FakeOrchestrator:
        def __init__(self, project_root):
            self.project_root = project_root
            self.loader = FakeLoader()
            self.config = TestbedConfig(root / "configs" / "testbed.yaml")

    monkeypatch.setattr(testpilot.cli, "Orchestrator", FakeOrchestrator)

    runner = CliRunner()
    result = runner.invoke(
        main,
        [
            "brcm-fw-upgrade",
            "run",
            "--case",
            "brcm-fw-upgrade-dut-sta-forward",
            "--forward-image",
            "images/0410.pkgtb",
            "--rollback-image",
            "images/0403.pkgtb",
            "--fw-name",
            "bcmBGW720-300_squashfs_full_update.pkgtb",
            "--expected-image-tag",
            "631BGW720-3001101323",
            "--expected-build-time",
            "Mon Apr 20 13:02:57 CST 2026",
            "--platform-profile",
            "bgw720_prpl",
            "--topology",
            "single_dut",
        ],
    )

    assert result.exit_code != 0
    assert "requires topology 'dut_plus_sta', got 'single_dut'" in result.output
    assert "Error:" in result.output


def test_brcm_fw_upgrade_run_rejects_unknown_case_with_clean_cli_error(monkeypatch):
    _clear_provider_env(monkeypatch)
    root = Path(__file__).resolve().parents[1]

    class FakeLoader:
        def load(self, plugin_name: str):
            assert plugin_name == "brcm_fw_upgrade"
            return _load_real_brcm_plugin()

    class FakeOrchestrator:
        def __init__(self, project_root):
            self.project_root = project_root
            self.loader = FakeLoader()
            self.config = TestbedConfig(root / "configs" / "testbed.yaml")

    monkeypatch.setattr(testpilot.cli, "Orchestrator", FakeOrchestrator)

    runner = CliRunner()
    result = runner.invoke(
        main,
        [
            "brcm-fw-upgrade",
            "run",
            "--case",
            "brcm-fw-upgrade-dut-sta-forwad",
            "--forward-image",
            "images/0410.pkgtb",
            "--rollback-image",
            "images/0403.pkgtb",
            "--fw-name",
            "bcmBGW720-300_squashfs_full_update.pkgtb",
            "--expected-image-tag",
            "631BGW720-3001101323",
            "--expected-build-time",
            "Mon Apr 20 13:02:57 CST 2026",
            "--platform-profile",
            "bgw720_prpl",
            "--topology",
            "dut_plus_sta",
        ],
    )

    assert result.exit_code != 0
    assert "Error: unknown case id(s): brcm-fw-upgrade-dut-sta-forwad" in result.output
    assert "Traceback" not in result.output


def test_baseline_qualify_invokes_wifi_llapi_plugin(monkeypatch):
    """wifi-llapi baseline-qualify dispatches to plugin.qualify_baseline."""
    _clear_provider_env(monkeypatch)
    calls: list[dict[str, Any]] = []

    class FakePlugin:
        def qualify_baseline(self, topology, *, bands, repeat_count, soak_minutes):
            calls.append(
                {
                    "topology": topology,
                    "bands": bands,
                    "repeat_count": repeat_count,
                    "soak_minutes": soak_minutes,
                }
            )
            return {
                "overall_status": "stable",
                "bands": [{"band": "5g", "stable": True}],
                "repeat_count": repeat_count,
                "soak_minutes": soak_minutes,
            }

    class FakeLoader:
        def __init__(self) -> None:
            self.plugin = FakePlugin()

        def load(self, plugin_name: str):
            assert plugin_name == "wifi_llapi"
            return self.plugin

    class FakeOrchestrator:
        def __init__(self, project_root):
            self.project_root = project_root
            self.loader = FakeLoader()
            self.config = {"name": "fake-topology"}

    monkeypatch.setattr(testpilot.cli, "Orchestrator", FakeOrchestrator)

    runner = CliRunner()
    result = runner.invoke(
        main,
        [
            "wifi-llapi",
            "baseline-qualify",
            "--band",
            "5g",
            "--repeat-count",
            "1",
            "--soak-minutes",
            "0",
        ],
    )

    assert result.exit_code == 0
    assert calls == [
        {
            "topology": {"name": "fake-topology"},
            "bands": ("5g",),
            "repeat_count": 1,
            "soak_minutes": 0,
        }
    ]
    assert "overall_status" in result.output
    assert "stable" in result.output


def test_help_text_for_run():
    """run --help shows options."""
    runner = CliRunner()
    result = runner.invoke(main, ["run", "--help"])
    assert result.exit_code == 0
    assert "PLUGIN_NAME" in result.output
    assert "--case" in result.output
    assert "--dut-fw-ver" in result.output
    assert "Correct format:" in result.output
    assert "testpilot run wifi_llapi --case wifi-llapi-D004-kickstation" in result.output


# --- Plugin-scoped testbed staging --------------------------------------------


def test_list_cases_stages_plugin_testbed_template(tmp_path: Path, monkeypatch) -> None:
    """list-cases <plugin> copies plugins/<plugin>/testbed.yaml.example into configs/testbed.yaml."""
    _clear_provider_env(monkeypatch)
    _make_stage_marker_plugin(tmp_path, "stage_marker", "stage-marker-bench")

    runner = CliRunner()
    result = runner.invoke(main, ["--root", str(tmp_path), "list-cases", "stage_marker"])

    assert result.exit_code == 0, result.output
    staged = tmp_path / "configs" / "testbed.yaml"
    assert staged.exists()
    assert "stage-marker-bench" in staged.read_text(encoding="utf-8")


def test_run_switches_testbed_between_plugins(tmp_path: Path, monkeypatch) -> None:
    """Running plugin A then plugin B must overwrite configs/testbed.yaml each time."""
    _clear_provider_env(monkeypatch)
    _make_stage_marker_plugin(tmp_path, "plugin_a", "marker-A")
    _make_stage_marker_plugin(tmp_path, "plugin_b", "marker-B")

    runner = CliRunner()
    res_a = runner.invoke(main, ["--root", str(tmp_path), "list-cases", "plugin_a"])
    assert res_a.exit_code == 0, res_a.output
    staged = tmp_path / "configs" / "testbed.yaml"
    assert "marker-A" in staged.read_text(encoding="utf-8")

    res_b = runner.invoke(main, ["--root", str(tmp_path), "list-cases", "plugin_b"])
    assert res_b.exit_code == 0, res_b.output
    assert "marker-B" in staged.read_text(encoding="utf-8")
    assert "marker-A" not in staged.read_text(encoding="utf-8")


def test_list_plugins_does_not_overwrite_local_testbed(tmp_path: Path, monkeypatch) -> None:
    """list-plugins is plugin-agnostic and must not stage any plugin's testbed."""
    _clear_provider_env(monkeypatch)
    _make_stage_marker_plugin(tmp_path, "stage_marker", "stage-marker-bench")
    pre_existing = tmp_path / "configs" / "testbed.yaml"
    pre_existing.write_text("testbed:\n  name: user-local\n", encoding="utf-8")

    runner = CliRunner()
    result = runner.invoke(main, ["--root", str(tmp_path), "list-plugins"])

    assert result.exit_code == 0, result.output
    assert pre_existing.read_text(encoding="utf-8") == "testbed:\n  name: user-local\n"


def test_brcm_fw_upgrade_run_stages_brcm_testbed_before_run(monkeypatch, tmp_path: Path) -> None:
    """brcm-fw-upgrade run must stage plugins/brcm_fw_upgrade/testbed.yaml.example."""
    _clear_provider_env(monkeypatch)

    forward_image = tmp_path / "fw.pkgtb"
    forward_image.write_bytes(b"$imageversion: tag$ #1 SMP PREEMPT Mon Apr 20 13:02:57 CST 2026")

    captured: list[tuple[Path, str, Path]] = []
    real_stage = testpilot.cli.stage_plugin_testbed  # will fail until CLI imports it

    def spy(plugins_dir: Path, plugin_name: str, configs_dir: Path) -> Path:
        captured.append((plugins_dir, plugin_name, configs_dir))
        return real_stage(plugins_dir, plugin_name, configs_dir)

    monkeypatch.setattr(testpilot.cli, "stage_plugin_testbed", spy)
    # Short-circuit the heavy run path — we only care about staging happening first.
    monkeypatch.setattr(
        "testpilot.cli._derive_brcm_image_metadata",
        lambda *_args, **_kwargs: {"image_tag": "tag", "build_time": "Apr 20 13:02:57 CST 2026"},
    )

    class _StubPlugin:
        def run_cases(self, *_args, **_kwargs):
            return {"status": "ok", "topology_name": "lab-bench-1"}

    def _stub_load(self, name):
        return _StubPlugin()

    monkeypatch.setattr(
        "testpilot.core.plugin_loader.PluginLoader.load",
        _stub_load,
    )

    runner = CliRunner()
    result = runner.invoke(
        main,
        [
            "brcm-fw-upgrade",
            "run",
            "--forward-image",
            str(forward_image),
        ],
    )

    assert result.exit_code == 0, result.output
    assert any(plugin == "brcm_fw_upgrade" for _, plugin, _ in captured), captured


def test_baseline_qualify_stages_wifi_llapi_testbed_before_qualify(monkeypatch) -> None:
    """wifi-llapi baseline-qualify must stage plugins/wifi_llapi/testbed.yaml.example."""
    _clear_provider_env(monkeypatch)
    captured: list[tuple[Path, str, Path]] = []
    real_stage = testpilot.cli.stage_plugin_testbed

    def spy(plugins_dir: Path, plugin_name: str, configs_dir: Path) -> Path:
        captured.append((plugins_dir, plugin_name, configs_dir))
        return real_stage(plugins_dir, plugin_name, configs_dir)

    monkeypatch.setattr(testpilot.cli, "stage_plugin_testbed", spy)

    class _StubPlugin:
        def qualify_baseline(self, *_args, **_kwargs):
            return {"overall_status": "stable"}

    monkeypatch.setattr(
        "testpilot.core.plugin_loader.PluginLoader.load",
        lambda self, name: _StubPlugin(),
    )

    runner = CliRunner()
    result = runner.invoke(main, ["wifi-llapi", "baseline-qualify", "--band", "5g", "--repeat-count", "1", "--soak-minutes", "0"])

    assert result.exit_code == 0, result.output
    assert any(plugin == "wifi_llapi" for _, plugin, _ in captured), captured


def _make_cli_template_xlsx(path: Path) -> None:
    """Minimal template compatible with Task 2 validator (Summary + Wifi_LLAPI sheets)."""
    wb = Workbook()
    default_sheet = wb.active
    wb.remove(default_sheet)

    ws_sum = wb.create_sheet("Summary")
    headers = [
        "Module", "Object Category", "Total Items", "Tested Items",
        "Pass", "Fail", "To be confirmed", "Not Supported", "Skip",
        "Pass Rate", "result empty", "Progress",
    ]
    for col_idx, header in enumerate(headers, start=1):
        ws_sum.cell(row=2, column=col_idx).value = header
    ws_sum["B3"] = "WiFi.AccessPoint"
    ws_sum["F3"] = '=COUNTIFS(Wifi_LLAPI!$A:$A,$B3&"*",Wifi_LLAPI!$N:$N,F$2)'
    ws_sum["J3"] = "=IFERROR(E3/SUM(E3:F3),0)"
    ws_sum["B9"] = "WiFi.AccessPoint"
    ws_sum["F9"] = '=COUNTIFS(Wifi_LLAPI!$A:$A,$B9&"*",Wifi_LLAPI!$O:$O,F$2)'
    ws_sum["J9"] = "=IFERROR(E9/SUM(E9:F9),0)"
    ws_sum["B15"] = "WiFi.AccessPoint"
    ws_sum["F15"] = '=COUNTIFS(Wifi_LLAPI!$A:$A,$B15&"*",Wifi_LLAPI!$P:$P,F$2)'
    ws_sum["J15"] = "=IFERROR(E15/SUM(E15:F15),0)"
    for row in range(3, 21):
        ws_sum[f"J{row}"] = f"=IFERROR(E{row}/SUM(E{row}:F{row}),0)"

    ws = wb.create_sheet("Wifi_LLAPI")
    ws["A1"] = "Object"
    ws["E1"] = "LLAPI test steps"
    ws["L1"] = "Tester"
    ws["M1"] = "Comment"
    ws.merge_cells(start_row=2, end_row=2, start_column=9, end_column=11)
    ws.cell(row=2, column=9).value = "Result"
    ws["I3"] = "WiFi 5G"
    ws["J3"] = "WiFi 6G"
    ws["K3"] = "WiFi 2.4G"
    ws["L3"] = "Tester"
    ws["M3"] = "Comment"
    ws["N3"] = "Summary Bucket WiFi 5G"
    ws["O3"] = "Summary Bucket WiFi 6G"
    ws["P3"] = "Summary Bucket WiFi 2.4G"
    ws.column_dimensions["N"].hidden = True
    ws.column_dimensions["O"].hidden = True
    ws.column_dimensions["P"].hidden = True
    ws["A4"] = "WiFi.AccessPoint.1."
    ws["C4"] = "getSomeAPI"
    wb.save(path)


def test_wifi_llapi_reproject_summary_cli(tmp_path: Path, monkeypatch) -> None:
    """wifi-llapi reproject-summary exits 0 and emits xlsx + md + html + json artifacts."""
    _clear_provider_env(monkeypatch)

    # Set up fake repo root with template in expected location
    template_dir = tmp_path / "plugins" / "wifi_llapi" / "reports" / "templates"
    template_dir.mkdir(parents=True)
    template_path = template_dir / "wifi_llapi_template.xlsx"
    _make_cli_template_xlsx(template_path)

    # Write a minimal source JSON with one wifi_llapi case
    source_json_path = tmp_path / "source_run.json"
    source_data: dict[str, Any] = {
        "meta": {"title": "cli test run", "plugin": "wifi_llapi"},
        "cases": [
            {
                "case_id": "D001",
                "source_row": 4,
                "executed_test_command": "wl status",
                "command_output": "ok",
                "result_5g": "Pass",
                "result_6g": "Pass",
                "result_24g": "Pass",
                "diagnostic_status": "Pass",
                "comment": "",
            }
        ],
    }
    source_json_path.write_text(json.dumps(source_data), encoding="utf-8")

    out_dir = tmp_path / "out"

    runner = CliRunner()
    result = runner.invoke(
        main,
        [
            "--root", str(tmp_path),
            "wifi-llapi", "reproject-summary",
            "--source-json", str(source_json_path),
            "--out-dir", str(out_dir),
            "--output-stem", "out-report",
        ],
    )

    assert result.exit_code == 0, result.output
    assert (out_dir / "out-report.xlsx").exists(), result.output
    assert (out_dir / "out-report.md").exists(), result.output
    assert (out_dir / "out-report.html").exists(), result.output
    assert (out_dir / "out-report.json").exists(), result.output

    # stdout must be valid JSON with key fields
    clean = re.sub(r"\x1b\[[0-9;]*m", "", result.output)
    payload = json.loads(clean)
    assert payload["status"] == "ok"
    assert "report_path" in payload
    assert "summary" in payload
