"""Tests for offline wifi_llapi reproject orchestration (Task 3 + 4)."""

from __future__ import annotations

import json
from pathlib import Path
from typing import Any

import pytest
from openpyxl import Workbook, load_workbook

from testpilot.reporting.wifi_llapi_summary import (
    SUMMARY_POLICY_VERSION,
    extract_fail_reason,
)

# ---------------------------------------------------------------------------
# Synthetic data helpers
# ---------------------------------------------------------------------------

_D001: dict[str, Any] = {
    "case_id": "D001",
    "source_row": 4,
    "executed_test_command": "wl status",
    "command_output": "ok",
    "result_5g": "Pass",
    "result_6g": "Fail",
    "result_24g": "Not Supported",
    "diagnostic_status": "FailEnv",
    "comment": "env_verify gate failed",
    "failure_snapshot": {
        "phase": "verify_env",
        "reason_code": "sta_band_not_ready",
    },
}

_D002: dict[str, Any] = {
    "case_id": "D002",
    "source_row": 5,
    "executed_test_command": "wl assoclist",
    "command_output": "ok",
    "result_5g": "Fail",
    "result_6g": "Skip",
    "result_24g": "N/A",
    "diagnostic_status": "FailTest",
    "comment": "pass_criteria not satisfied",
    "failure_snapshot": {
        "phase": "evaluate",
        "reason_code": "criteria_mismatch",
    },
}

_SOURCE_JSON: dict[str, Any] = {
    "meta": {"title": "test run", "plugin": "wifi_llapi"},
    "cases": [_D001, _D002],
}


def _make_template_xlsx(path: Path) -> None:
    """Create a minimal synthetic template compatible with Task 2 validator."""
    wb = Workbook()
    default_sheet = wb.active
    wb.remove(default_sheet)

    # Summary sheet (required headers in row 3)
    ws_sum = wb.create_sheet("Summary")
    headers = [
        "Module", "Object Category", "Total Items", "Tested Items",
        "Pass", "Fail", "To be tested", "Not Supported", "Skip",
        "Pass Rate", "result empty", "Progress",
    ]
    for col_idx, header in enumerate(headers, start=1):
        ws_sum.cell(row=3, column=col_idx).value = header

    # Wifi_LLAPI sheet
    ws = wb.create_sheet("Wifi_LLAPI")
    ws["A1"] = "Object"
    ws["E1"] = "LLAPI test steps"
    ws["L1"] = "Tester"
    ws["M1"] = "Comment"
    # Row 2: Result group header (merged I~K)
    ws.merge_cells(start_row=2, end_row=2, start_column=9, end_column=11)
    ws.cell(row=2, column=9).value = "Result"
    ws.cell(row=2, column=12).value = "Tester"
    # Row 3: individual band headers
    ws["I3"] = "WiFi 5G"
    ws["J3"] = "WiFi 6G"
    ws["K3"] = "WiFi 2.4G"
    ws["L3"] = "Tester"
    ws["M3"] = "Comment"
    # Data rows (object + api columns)
    ws["A4"] = "WiFi.AccessPoint.1."
    ws["C4"] = "getSomeAPI"
    ws["A5"] = "WiFi.EndPoint.1."
    ws["C5"] = "getAnotherAPI"

    wb.save(path)
    wb.close()


# ---------------------------------------------------------------------------
# Task 3 – reproject tests
# ---------------------------------------------------------------------------


def test_reproject_creates_isolated_artifacts_and_preserves_source(
    tmp_path: Path,
) -> None:
    from testpilot.reporting.wifi_llapi_reproject import reproject_wifi_llapi_report

    template = tmp_path / "template.xlsx"
    _make_template_xlsx(template)

    source_json = tmp_path / "source.json"
    source_json.write_text(
        json.dumps(_SOURCE_JSON, ensure_ascii=False), encoding="utf-8"
    )
    original_text = source_json.read_text(encoding="utf-8")

    out_dir = tmp_path / "out"
    result = reproject_wifi_llapi_report(
        source_json=source_json,
        template_xlsx=template,
        out_dir=out_dir,
        output_stem="new-report",
    )

    # Source JSON must be unchanged
    assert source_json.read_text(encoding="utf-8") == original_text

    # All expected paths present
    assert result["artifact_dir"].is_dir()
    assert result["report_path"].is_file()
    assert result["md_report_path"].is_file()
    assert result["html_report_path"].is_file()
    assert result["json_report_path"].is_file()

    # XLSX: check I/J/K raw values and M fail-reason for each case
    wb = load_workbook(result["report_path"], data_only=True)
    ws = wb["Wifi_LLAPI"]

    # D001 at row 4
    assert str(ws.cell(row=4, column=9).value or "") == "Pass"   # I = result_5g
    assert str(ws.cell(row=4, column=10).value or "") == "Fail"  # J = result_6g
    assert str(ws.cell(row=4, column=11).value or "") == "Not Supported"  # K = result_24g
    # M = extract_fail_reason(D001): reason_code "sta_band_not_ready" → "sta band not ready"
    assert ws.cell(row=4, column=13).value == extract_fail_reason(_D001)

    # D002 at row 5
    assert str(ws.cell(row=5, column=9).value or "") == "Fail"   # I = result_5g
    assert str(ws.cell(row=5, column=10).value or "") == "Skip"  # J = result_6g
    assert str(ws.cell(row=5, column=11).value or "") == "N/A"   # K = result_24g
    assert ws.cell(row=5, column=13).value == extract_fail_reason(_D002)
    wb.close()

    # Summary sheet: B1 policy version, A4/B4 row shape
    wb2 = load_workbook(result["report_path"], data_only=True)
    ws_sum = wb2["Summary"]
    assert ws_sum.cell(row=1, column=2).value == SUMMARY_POLICY_VERSION
    # Row 4 is first band_category row (5G / WiFi.AccessPoint)
    assert ws_sum.cell(row=4, column=1).value is not None  # band label
    assert ws_sum.cell(row=4, column=2).value is not None  # category
    wb2.close()

    # JSON report: meta.source_json, summary policy_version, to_be_tested count
    report_data = json.loads(result["json_report_path"].read_text(encoding="utf-8"))
    assert report_data["meta"]["source_json"].endswith("source.json")
    summary = report_data["summary"]
    assert summary["policy_version"] == SUMMARY_POLICY_VERSION
    # D001 result_6g=Fail + FailEnv → "To be tested"; expect bucket_totals count
    assert summary["bucket_totals"]["result_6g"]["to_be_tested"] == 1


def test_reproject_raises_if_out_dir_non_empty(tmp_path: Path) -> None:
    from testpilot.reporting.wifi_llapi_reproject import reproject_wifi_llapi_report

    template = tmp_path / "template.xlsx"
    _make_template_xlsx(template)
    source_json = tmp_path / "source.json"
    source_json.write_text(json.dumps(_SOURCE_JSON, ensure_ascii=False), encoding="utf-8")

    out_dir = tmp_path / "existing"
    out_dir.mkdir()
    (out_dir / "existing_file.txt").write_text("occupied")

    with pytest.raises(FileExistsError):
        reproject_wifi_llapi_report(
            source_json=source_json,
            template_xlsx=template,
            out_dir=out_dir,
        )


def test_reproject_raises_on_non_dict_source_json(tmp_path: Path) -> None:
    """Source JSON that is an array (not an object) must raise TypeError/ValueError."""
    from testpilot.reporting.wifi_llapi_reproject import reproject_wifi_llapi_report

    template = tmp_path / "template.xlsx"
    _make_template_xlsx(template)

    # Write a JSON array instead of an object
    source_json = tmp_path / "bad_source.json"
    source_json.write_text(json.dumps([_D001, _D002]), encoding="utf-8")

    with pytest.raises((TypeError, ValueError), match="JSON"):
        reproject_wifi_llapi_report(
            source_json=source_json,
            template_xlsx=template,
            out_dir=tmp_path / "out_bad",
        )


def test_reproject_default_out_dir_anchored_to_template_parent(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    """Default out_dir must be under template_path.parent.parent, not CWD."""
    from testpilot.reporting.wifi_llapi_reproject import reproject_wifi_llapi_report

    # Place template in a subdirectory mirroring the real layout
    template_dir = tmp_path / "plugins" / "wifi_llapi" / "reports" / "templates"
    template_dir.mkdir(parents=True)
    template = template_dir / "template.xlsx"
    _make_template_xlsx(template)

    source_json = tmp_path / "source.json"
    source_json.write_text(json.dumps(_SOURCE_JSON, ensure_ascii=False), encoding="utf-8")

    # Change CWD to a different directory so CWD-relative paths would be wrong
    monkeypatch.chdir(tmp_path / "plugins")

    result = reproject_wifi_llapi_report(
        source_json=source_json,
        template_xlsx=template,
        # no out_dir — uses default
    )

    # artifact_dir.parent must be template_path.parent.parent
    # i.e. plugins/wifi_llapi/reports (sibling of templates/)
    expected_parent = template.parent.parent.resolve()
    assert result["artifact_dir"].resolve().parent == expected_parent
