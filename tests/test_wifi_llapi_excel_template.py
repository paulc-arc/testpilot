from datetime import date
from pathlib import Path

from openpyxl import Workbook, load_workbook

from testpilot.reporting.wifi_llapi_excel import (
    WifiLlapiCaseResult,
    build_template_from_source,
    collect_alignment_issues,
    create_run_report_from_template,
    fill_case_results,
    generate_report_filename,
)


def _create_source_xlsx(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Wifi_LLAPI"
    wb.create_sheet("OtherSheet")

    ws["A2"] = "Object"
    ws["C2"] = "Parameter Name"
    ws["G2"] = "Test steps"
    ws["H2"] = "Command Output"
    ws["S2"] = "BCM 5g result"
    ws["T2"] = "BCM 6g result"
    ws["U2"] = "BCM 2.4g result"
    ws["V2"] = "BCM Comment"

    ws["A4"] = "WiFi.AccessPoint.{i}."
    ws["C4"] = "kickStation()"
    ws["G4"] = "old-step"
    ws["H4"] = "old-output"
    ws["S4"] = "Pass"
    ws["T4"] = "Pass"
    ws["U4"] = "Pass"
    ws["V4"] = "old-comment"

    ws["A5"] = "WiFi.Radio.{i}."
    ws["C5"] = "scan()"
    ws["G5"] = "old-step2"
    ws["H5"] = "old-output2"
    ws["S5"] = "Fail"
    ws["T5"] = "Fail"
    ws["U5"] = "Fail"
    ws["V5"] = "old-comment2"

    wb.save(path)
    wb.close()


def test_build_template_from_source(tmp_path: Path):
    source = tmp_path / "source.xlsx"
    template = tmp_path / "wifi_llapi_template.xlsx"
    _create_source_xlsx(source)

    result = build_template_from_source(source, template)
    assert result.total_case_rows == 2
    assert result.sheet_name == "Wifi_LLAPI"

    wb = load_workbook(template)
    assert wb.sheetnames == ["Wifi_LLAPI"]
    ws = wb["Wifi_LLAPI"]
    assert ws["C4"].value == "kickStation()"
    assert ws["C5"].value == "scan()"
    for cell in ("G4", "H4", "S4", "T4", "U4", "V4", "G5", "H5", "S5", "T5", "U5", "V5"):
        assert ws[cell].value is None
    wb.close()


def test_fill_case_results(tmp_path: Path):
    source = tmp_path / "source.xlsx"
    template = tmp_path / "wifi_llapi_template.xlsx"
    report = tmp_path / "20260304_BGW720-B0-403_wifi_LLAPI.xlsx"
    _create_source_xlsx(source)
    build_template_from_source(source, template)
    create_run_report_from_template(template, report)

    fill_case_results(
        report,
        [
            WifiLlapiCaseResult(
                case_id="wifi-llapi-r006-kickstation",
                source_row=4,
                executed_test_command="ubus-cli ...kickStation...",
                command_output="assoclist empty",
                result_5g="Pass",
                result_6g="Pass",
                result_24g="Pass",
                comment="ok",
            )
        ],
    )

    wb = load_workbook(report)
    ws = wb["Wifi_LLAPI"]
    assert ws["G4"].value == "ubus-cli ...kickStation..."
    assert ws["H4"].value == "assoclist empty"
    assert ws["S4"].value == "Pass"
    assert ws["T4"].value == "Pass"
    assert ws["U4"].value == "Pass"
    assert ws["V4"].value == "ok"
    wb.close()


def test_generate_report_filename():
    assert (
        generate_report_filename(date(2026, 3, 4), "BGW720-B0-403")
        == "20260304_BGW720-B0-403_wifi_LLAPI.xlsx"
    )
    assert (
        generate_report_filename(date(2026, 3, 4), "BGW720 B0 403")
        == "20260304_BGW720-B0-403_wifi_LLAPI.xlsx"
    )


def test_collect_alignment_issues(tmp_path: Path):
    source = tmp_path / "source.xlsx"
    _create_source_xlsx(source)
    cases = [
        {
            "id": "wifi-llapi-kickstation",
            "source": {
                "row": 4,
                "object": "WiFi.AccessPoint.{i}.",
                "api": "kickStation()",
            },
        },
        {
            "id": "wifi-llapi-mismatch",
            "source": {
                "row": 5,
                "object": "WiFi.Radio.{i}.",
                "api": "wrongApi()",
            },
        },
    ]
    issues = collect_alignment_issues(cases, source)
    assert len(issues) == 1
    assert issues[0]["case_id"] == "wifi-llapi-mismatch"
    assert issues[0]["issue"] == "object_or_api_mismatch"
