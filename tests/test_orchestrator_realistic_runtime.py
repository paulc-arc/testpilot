"""Realistic integration tests for orchestrator wifi_llapi runtime path."""

from __future__ import annotations

import copy
import json
from pathlib import Path
import shutil
from typing import Any

from openpyxl import Workbook, load_workbook

from testpilot.core.orchestrator import Orchestrator

FAIL_CASE_ID = "wifi-llapi-r900-evaluate-fail"
PASS_CASE_ID = "wifi-llapi-r901-evaluate-pass"
PASS_TOKEN = "TOKEN_OK"


class MockTransport:
    """Test-only transport mock to simulate command execution."""

    def __init__(self) -> None:
        self._connected = False
        self.history: list[str] = []

    @property
    def is_connected(self) -> bool:
        return self._connected

    def connect(self) -> None:
        self._connected = True

    def disconnect(self) -> None:
        self._connected = False

    def execute(self, command: str, timeout: float = 30.0) -> dict[str, Any]:
        self.history.append(command)
        return {
            "returncode": 0,
            "stdout": f"[mock] timeout={timeout} {command}",
            "stderr": "",
            "elapsed": 0.01,
        }


def _write_testbed_yaml(path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(
        """
testbed:
  name: integration-testbed
  devices:
    DUT:
      role: ap
      transport: serial
    STA:
      role: sta
      transport: adb
""".strip()
        + "\n",
        encoding="utf-8",
    )


def _write_source_xlsx(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Wifi_LLAPI"

    ws["A2"] = "Object"
    ws["C2"] = "Parameter Name"
    ws["G2"] = "Test steps"
    ws["H2"] = "Command Output"
    ws["I2"] = "ARC 4.0.3 Test Result\nWiFi 5g"
    ws["J2"] = "ARC 4.0.3 Test Result\nWiFi 6g"
    ws["K2"] = "ARC 4.0.3 Test Result\nWiFi 2.4g"
    ws["L2"] = "Tester"
    ws["M2"] = "ARC Comment"

    ws["A4"] = "WiFi.AccessPoint.{i}."
    ws["C4"] = "kickStation()"
    ws["A5"] = "WiFi.Radio.{i}."
    ws["C5"] = "getRadioStats()"

    wb.save(path)
    wb.close()


def _prepare_runtime_project(tmp_path: Path) -> tuple[Path, Path]:
    project_root = tmp_path / "project"
    plugin_dir = project_root / "plugins" / "wifi_llapi"
    plugin_dir.mkdir(parents=True, exist_ok=True)

    repo_root = Path(__file__).resolve().parents[1]
    source_plugin_dir = repo_root / "plugins" / "wifi_llapi"
    shutil.copy2(source_plugin_dir / "plugin.py", plugin_dir / "plugin.py")
    shutil.copy2(source_plugin_dir / "agent-config.yaml", plugin_dir / "agent-config.yaml")

    _write_testbed_yaml(project_root / "configs" / "testbed.yaml")
    source_xlsx = project_root / "source.xlsx"
    _write_source_xlsx(source_xlsx)
    return project_root, source_xlsx


def _build_cases() -> list[dict[str, Any]]:
    return [
        {
            "id": FAIL_CASE_ID,
            "name": "evaluate-fail",
            "source": {
                "row": 4,
                "object": "WiFi.AccessPoint.{i}.",
                "api": "kickStation()",
            },
            "steps": [
                {
                    "id": "step1",
                    "target": "DUT",
                    "command": 'ubus-cli "WiFi.AccessPoint.1.kickStation(macaddress=AA:BB:CC:DD:EE:FF)"',
                }
            ],
            "pass_criteria": [
                {
                    "field": "step1.output",
                    "operator": "contains",
                    "value": PASS_TOKEN,
                }
            ],
            "emit_pass_token": False,
            "bands": ["5g"],
        },
        {
            "id": PASS_CASE_ID,
            "name": "evaluate-pass",
            "source": {
                "row": 5,
                "object": "WiFi.Radio.{i}.",
                "api": "getRadioStats()",
            },
            "steps": [
                {
                    "id": "step1",
                    "target": "STA",
                    "command": 'ubus-cli "WiFi.Radio.1.getRadioStats()"',
                }
            ],
            "pass_criteria": [
                {
                    "field": "step1.output",
                    "operator": "contains",
                    "value": PASS_TOKEN,
                }
            ],
            "emit_pass_token": True,
            "bands": ["6g"],
        },
    ]


def _patch_runtime_hooks(monkeypatch, plugin: Any, cases: list[dict[str, Any]]) -> dict[str, Any]:
    transports = {
        "DUT": MockTransport(),
        "STA": MockTransport(),
    }
    state: dict[str, Any] = {
        "setup_calls": [],
        "verify_calls": [],
        "execute_calls": [],
        "evaluate_calls": [],
        "teardown_calls": [],
        "transports": transports,
    }

    def discover_cases() -> list[dict[str, Any]]:
        return copy.deepcopy(cases)

    def setup_env(case: dict[str, Any], topology: Any) -> bool:
        state["setup_calls"].append(str(case.get("id", "?")))
        assert "DUT" in topology.devices
        assert "STA" in topology.devices
        for transport in transports.values():
            if not transport.is_connected:
                transport.connect()
        return True

    def verify_env(case: dict[str, Any], topology: Any) -> bool:
        case_id = str(case.get("id", "?"))
        state["verify_calls"].append(case_id)
        probe = transports["DUT"].execute(f"verify:{case_id}", timeout=5.0)
        return probe["returncode"] == 0

    def execute_step(case: dict[str, Any], step: dict[str, Any], topology: Any) -> dict[str, Any]:
        case_id = str(case.get("id", "?"))
        step_id = str(step.get("id", "step"))
        attempt = int(case.get("_attempt_index", 0))
        target = str(step.get("target", "DUT"))
        timeout = float(step.get("timeout", 0.0))
        command = str(step.get("command", ""))
        transport = transports[target]

        result = transport.execute(command, timeout=timeout)
        output = str(result.get("stdout", ""))
        if bool(case.get("emit_pass_token")):
            output = f"{output} {PASS_TOKEN}"
        else:
            output = f"{output} TOKEN_MISSING"

        state["execute_calls"].append(
            {
                "case_id": case_id,
                "attempt": attempt,
                "step_id": step_id,
                "target": target,
                "timeout": timeout,
            }
        )
        return {
            "success": True,
            "output": output,
            "captured": {"stdout": result.get("stdout", "")},
            "timing": float(result.get("elapsed", 0.0)),
        }

    def evaluate(case: dict[str, Any], results: dict[str, Any]) -> bool:
        case_id = str(case.get("id", "?"))
        attempt = int(case.get("_attempt_index", 0))
        state["evaluate_calls"].append({"case_id": case_id, "attempt": attempt})

        steps = results.get("steps", {})
        if not isinstance(steps, dict):
            return False

        for criterion in case.get("pass_criteria", []):
            if not isinstance(criterion, dict):
                continue
            field = str(criterion.get("field", ""))
            operator = str(criterion.get("operator", ""))
            expected = str(criterion.get("value", ""))

            step_key, _, attr = field.partition(".")
            step_result = steps.get(step_key, {})
            if not isinstance(step_result, dict):
                return False
            actual = str(step_result.get(attr, ""))
            if operator == "contains" and expected not in actual:
                return False
        return True

    def teardown(case: dict[str, Any], topology: Any) -> None:
        state["teardown_calls"].append(str(case.get("id", "?")))
        for transport in transports.values():
            if transport.is_connected:
                transport.disconnect()

    monkeypatch.setattr(plugin, "discover_cases", discover_cases)
    monkeypatch.setattr(plugin, "setup_env", setup_env)
    monkeypatch.setattr(plugin, "verify_env", verify_env)
    monkeypatch.setattr(plugin, "execute_step", execute_step)
    monkeypatch.setattr(plugin, "evaluate", evaluate)
    monkeypatch.setattr(plugin, "teardown", teardown)
    return state


def _load_case_traces(trace_dir: Path) -> dict[str, dict[str, Any]]:
    traces: dict[str, dict[str, Any]] = {}
    for path in sorted(trace_dir.glob("*.json")):
        payload = json.loads(path.read_text(encoding="utf-8"))
        case_id = payload.get("case_id")
        if isinstance(case_id, str):
            traces[case_id] = payload
    return traces


def _run_realistic_runtime(tmp_path: Path, monkeypatch) -> tuple[dict[str, Any], dict[str, Any]]:
    project_root, source_xlsx = _prepare_runtime_project(tmp_path)
    orch = Orchestrator(
        project_root=project_root,
        plugins_dir=project_root / "plugins",
        config_path=project_root / "configs" / "testbed.yaml",
    )

    plugin = orch.loader.load("wifi_llapi")
    cases = _build_cases()
    state = _patch_runtime_hooks(monkeypatch, plugin=plugin, cases=cases)

    result = orch.run(
        "wifi_llapi",
        case_ids=[FAIL_CASE_ID, PASS_CASE_ID],
        dut_fw_ver="FW-IT-REALISTIC-1",
        report_source_xlsx=str(source_xlsx),
    )
    return result, state


def test_realistic_runtime_covers_hooks_and_report_outputs(tmp_path: Path, monkeypatch):
    result, state = _run_realistic_runtime(tmp_path, monkeypatch)

    assert result["status"] == "completed"
    assert result["cases_count"] == 2
    assert result["pass_count"] == 1
    assert result["fail_count"] == 1
    assert result["agent_trace_count"] == 2

    report_path = Path(result["report_path"])
    trace_dir = Path(result["agent_trace_dir"])
    assert report_path.is_file()
    assert trace_dir.is_dir()

    # 1 fail case (2 attempts) + 1 pass case (1 attempt) = 3 executions per hook.
    assert len(state["setup_calls"]) == 3
    assert len(state["verify_calls"]) == 3
    assert len(state["execute_calls"]) == 3
    assert len(state["evaluate_calls"]) == 3
    assert len(state["teardown_calls"]) == 3

    wb = load_workbook(report_path)
    ws = wb["Wifi_LLAPI"]
    assert ws["G4"].value == 'ubus-cli "WiFi.AccessPoint.1.kickStation(macaddress=AA:BB:CC:DD:EE:FF)"'
    assert ws["I4"].value == "Fail"
    assert ws["J4"].value == "N/A"
    assert ws["K4"].value == "N/A"
    assert ws["L4"].value == "testpilot"
    assert ws.max_column == 12

    assert ws["G5"].value == 'ubus-cli "WiFi.Radio.1.getRadioStats()"'
    assert PASS_TOKEN in str(ws["H5"].value)
    assert ws["I5"].value == "N/A"
    assert ws["J5"].value == "Pass"
    assert ws["K5"].value == "N/A"
    assert ws["L5"].value == "testpilot"

    meta = wb["_meta"]
    assert meta["B2"].value == "FW-IT-REALISTIC-1"
    wb.close()

    transports = state["transports"]
    assert any(cmd.startswith("verify:") for cmd in transports["DUT"].history)
    assert any("kickStation" in cmd for cmd in transports["DUT"].history)
    assert any("getRadioStats" in cmd for cmd in transports["STA"].history)


def test_fail_and_continue_with_plugin_evaluate_failure(tmp_path: Path, monkeypatch):
    result, state = _run_realistic_runtime(tmp_path, monkeypatch)

    assert result["status"] == "completed"
    assert result["pass_count"] == 1
    assert result["fail_count"] == 1

    execute_case_order = [item["case_id"] for item in state["execute_calls"]]
    assert execute_case_order == [FAIL_CASE_ID, FAIL_CASE_ID, PASS_CASE_ID]

    evaluate_case_order = [item["case_id"] for item in state["evaluate_calls"]]
    assert evaluate_case_order == [FAIL_CASE_ID, FAIL_CASE_ID, PASS_CASE_ID]

    traces = _load_case_traces(Path(result["agent_trace_dir"]))
    assert set(traces.keys()) == {FAIL_CASE_ID, PASS_CASE_ID}

    fail_trace = traces[FAIL_CASE_ID]
    pass_trace = traces[PASS_CASE_ID]

    assert fail_trace["final"]["status"] == "Fail"
    assert fail_trace["final"]["attempts_used"] == 2
    assert len(fail_trace["attempts"]) == 2
    assert all(item["status"] == "Fail" for item in fail_trace["attempts"])
    assert "pass_criteria not satisfied" in str(fail_trace["final"]["comment"])

    assert pass_trace["final"]["status"] == "Pass"
    assert pass_trace["final"]["attempts_used"] == 1
    assert len(pass_trace["attempts"]) == 1
    assert pass_trace["attempts"][0]["status"] == "Pass"
